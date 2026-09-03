"""
JX3Manager - 剑网3多角色管理器 (PyQt6 现代化桌面版)
"""
import os
import sys
import json
import csv
import re
import datetime
from typing import List, Dict, Any, Tuple, Optional
from PIL import Image

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QComboBox, QLineEdit, QTabWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QScrollArea, QFrame, QDialog, QMenu,
    QFileDialog, QMessageBox, QToolTip, QGridLayout, QSizePolicy, QCheckBox, QDoubleSpinBox, QGroupBox, QStackedWidget, QInputDialog, QTextEdit,
    QAbstractItemView, QListWidget, QListWidgetItem, QSpinBox, QSplitter
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize, QSettings, QUrl
from PyQt6.QtGui import QFont, QColor, QPalette, QIcon, QPixmap, QCursor, QBrush, QDesktopServices

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from main import JX3Manager, get_boss_aliases, get_floors_for_skill_boss, ANOMALY_MAP, filter_cd_dungeon_ids, filter_out_benched
from readers.plugin_settings import enable_all_stats
from combat_log_config import enable_combat_logs_for_all
from config_loader import get_cached_config, save_config, validate_config
from path_detector import detect_game_path, is_valid_game_path
from logger import get_logger

logger = get_logger("gui_qt")

APP_ICON_PATH = os.path.join(os.path.dirname(__file__), "resources", "app_icon.png")

def get_app_icon():
    """获取应用图标，文件不存在时返回空的 QIcon"""
    if os.path.exists(APP_ICON_PATH):
        return QIcon(APP_ICON_PATH)
    return QIcon()

# 百战"全清"判定阈值：击杀数达到 12 即视为全清（本周排班满轮换 12 个首领）
BAIZHAN_CLEARED_THRESHOLD = 12

# 颜色映射表
COLOR_MAP = {
    "黄": "#ffcc00", "蓝": "#4488ff", "绿": "#44cc44",
    "红": "#ff4444", "紫": "#bb44ff", "黑": "#555555", "白": "#dddddd"
}

COLOR_ORDER = {
    2: (1, "黄破绽"),
    3: (2, "蓝破绽"),
    4: (3, "绿破绽"),
    5: (4, "红破绽"),
    6: (5, "紫破绽"),
    7: (6, "黑破绽"),
    0: (7, "白破绽")
}

COLOR_STYLES = {
    2: ("#332b00", "#ffcc00", "#ffeb3b", "#ffcc00"),  # 黄
    3: ("#001a33", "#4488ff", "#90caf9", "#4488ff"),  # 蓝
    4: ("#002b00", "#44cc44", "#a5d6a7", "#44cc44"),  # 绿
    5: ("#330000", "#ff4444", "#ef9a9a", "#ff4444"),  # 红
    6: ("#260033", "#bb44ff", "#e1bee7", "#bb44ff"),  # 紫
    7: ("#1a1a1a", "#757575", "#e0e0e0", "#9e9e9e"),  # 黑
    0: ("#2a2a2a", "#cccccc", "#ffffff", "#e0e0e0")   # 白
}

TYPE_MAP = {
    "1": "攻击", "2": "攻击", "3": "攻击", "4": "攻击", "5": "攻击",
    "6": "控制", "7": "控制", "8": "控制", "9": "控制",
    "10": "位移", "11": "治疗", "12": "特殊", "13": "位移", "14": "位移"
}

DARK_QSS = """
QMainWindow {
    background-color: #121222;
}
QWidget {
    font-family: "Microsoft YaHei UI", "Segoe UI", sans-serif;
    color: #e0e0e0;
}
QFrame#HeaderFrame {
    background-color: #1a1a2e;
    border-bottom: 1px solid #2a2a4a;
}
QFrame#ToolbarFrame {
    background-color: #1a1a2e;
    border-radius: 8px;
}
QPushButton {
    background-color: #2b2b48;
    color: #ffffff;
    border: 1px solid #3b3b68;
    border-radius: 5px;
    padding: 6px 14px;
    font-size: 12px;
}
QPushButton:hover {
    background-color: #3b3b68;
    border-color: #4b4b88;
}
QPushButton:pressed {
    background-color: #1e1e38;
}
QPushButton#PrimaryBtn {
    background-color: #2a52be;
    border-color: #3b63ce;
}
QPushButton#PrimaryBtn:hover {
    background-color: #3b63ce;
}
QPushButton#GreenBtn {
    background-color: #2d5a2d;
    border-color: #3a6a3a;
}
QPushButton#GreenBtn:hover {
    background-color: #3a7a3a;
}
QPushButton#PurpleBtn {
    background-color: #4a2d5a;
    border-color: #6a3a7a;
}
QPushButton#PurpleBtn:hover {
    background-color: #6a3a8a;
}
QLineEdit, QSpinBox, QDoubleSpinBox {
    background-color: #232342;
    border: 1px solid #4a4a7a;
    border-radius: 5px;
    padding: 4px 8px;
    color: #ffffff;
    font-size: 12px;
    font-weight: bold;
    selection-background-color: #3b63ce;
    selection-color: #ffffff;
}
QLineEdit:hover, QSpinBox:hover, QDoubleSpinBox:hover {
    background-color: #2b2b52;
    border-color: #64b5f6;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
    border-color: #64b5f6;
    background-color: #28284c;
}
QLineEdit:disabled, QSpinBox:disabled, QDoubleSpinBox:disabled {
    background-color: #16162a;
    border-color: #2a2a4a;
    color: #666688;
}
QSpinBox::up-button, QDoubleSpinBox::up-button {
    subcontrol-origin: border;
    subcontrol-position: top right;
    width: 18px;
    border-left: 1px solid #3b3b68;
    border-bottom: 1px solid #3b3b68;
    border-top-right-radius: 4px;
    background-color: #232342;
}
QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover {
    background-color: #3b3b68;
}
QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {
    image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' viewBox='0 0 24 24' fill='none' stroke='%2364b5f6' stroke-width='3' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='18 15 12 9 6 15'%3E%3C/polyline%3E%3C/svg%3E");
    width: 10px;
    height: 10px;
}
QSpinBox::down-button, QDoubleSpinBox::down-button {
    subcontrol-origin: border;
    subcontrol-position: bottom right;
    width: 18px;
    border-left: 1px solid #3b3b68;
    border-bottom-right-radius: 4px;
    background-color: #232342;
}
QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover {
    background-color: #3b3b68;
}
QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {
    image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' viewBox='0 0 24 24' fill='none' stroke='%2364b5f6' stroke-width='3' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'%3E%3C/polyline%3E%3C/svg%3E");
    width: 10px;
    height: 10px;
}
QComboBox {
    background-color: #232342;
    border: 1px solid #4a4a7a;
    border-radius: 5px;
    padding: 5px 10px;
    color: #ffffff;
    font-size: 12px;
    font-weight: bold;
    min-height: 22px;
}
QComboBox:hover {
    background-color: #2b2b52;
    border-color: #64b5f6;
}
QComboBox:focus {
    border-color: #64b5f6;
}
QComboBox:disabled {
    background-color: #16162a;
    border-color: #2a2a4a;
    color: #666688;
}
QComboBox::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 24px;
    border-left: 1px solid #3b3b68;
    border-top-right-radius: 5px;
    border-bottom-right-radius: 5px;
    background-color: #232342;
}
QComboBox::drop-down:hover {
    background-color: #3b3b68;
}
QComboBox::down-arrow {
    image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%2364b5f6' stroke-width='3' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'%3E%3C/polyline%3E%3C/svg%3E");
    width: 12px;
    height: 12px;
    margin-right: 6px;
}
QComboBox QAbstractItemView {
    background-color: #1a1a32;
    border: 1px solid #4a4a7a;
    border-radius: 4px;
    color: #ffffff;
    selection-background-color: #3b63ce;
    selection-color: #ffffff;
    outline: 0px;
    padding: 4px;
}
QComboBox QAbstractItemView::item {
    min-height: 26px;
    padding: 4px 8px;
    color: #ffffff;
    background-color: #1a1a32;
}
QComboBox QAbstractItemView::item:hover,
QComboBox QAbstractItemView::item:selected {
    background-color: #3b63ce;
    color: #ffffff;
    font-weight: bold;
}
QTabWidget::pane {
    border: 1px solid #2a2a4a;
    background-color: #16162a;
    border-radius: 6px;
}
QTabBar::tab {
    background-color: #1e1e38;
    border: 1px solid #2a2a4a;
    padding: 8px 18px;
    margin-right: 4px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
}
QTabBar::tab:selected {
    background-color: #2b2b48;
    border-bottom-color: #2b2b48;
    color: #3b8ed0;
    font-weight: bold;
}
QTableWidget {
    background-color: #16162a;
    alternate-background-color: #1c1c34;
    gridline-color: #2a2a4a;
    border: none;
}
QHeaderView::section {
    background-color: #1e1e38;
    color: #a0a0c0;
    padding: 6px;
    border: 1px solid #2a2a4a;
    font-weight: bold;
}
QScrollArea {
    border: none;
    background-color: transparent;
}
QScrollBar:vertical {
    background: #16162a;
    width: 10px;
    margin: 0px;
}
QScrollBar::handle:vertical {
    background: #3b3b68;
    min-height: 20px;
    border-radius: 5px;
}
QScrollBar::handle:vertical:hover {
    background: #4b4b88;
}
"""

class NumericTableWidgetItem(QTableWidgetItem):
    """支持数值大小正常比较排序的 TableWidgetItem"""
    def __init__(self, display_text, numeric_val):
        super().__init__(str(display_text))
        self.numeric_val = numeric_val

    def __lt__(self, other):
        if isinstance(other, NumericTableWidgetItem):
            try:
                return self.numeric_val < other.numeric_val
            except TypeError:
                return str(self.numeric_val) < str(other.numeric_val)
        return super().__lt__(other)


class ConfigDialog(QDialog):
    """首次配置或缺少路径时的引导弹窗"""
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("剑3小助手 - 基础配置")
        self.resize(540, 290)
        self.setStyleSheet(DARK_QSS)
        icon = get_app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        title = QLabel("请设置剑网3客户端路径及 API Token：")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #3b8ed0;")
        layout.addWidget(title)

        # Game Path
        layout.addWidget(QLabel("游戏根目录 (例如 D:\\JX3_Classic):"))
        tip_path = QLabel("提示：应填到 ...\\bin\\zhcn_hd\\interface 层（即包含 my#data 的目录）")
        tip_path.setStyleSheet("font-size: 11px; color: #8888aa;")
        layout.addWidget(tip_path)

        path_layout = QHBoxLayout()
        path_layout.setSpacing(8)
        self.path_input = QLineEdit(self.config.get("game_path", ""))
        self.path_input.textChanged.connect(self.on_path_changed)
        btn_detect = QPushButton("🔍 自动检测")
        btn_detect.clicked.connect(self.run_detection)
        btn_browse = QPushButton("浏览...")
        btn_browse.clicked.connect(self.browse_path)
        path_layout.addWidget(self.path_input)
        path_layout.addWidget(btn_detect)
        path_layout.addWidget(btn_browse)
        layout.addLayout(path_layout)

        # 检测状态行
        self.lbl_detect_status = QLabel("")
        self.lbl_detect_status.setStyleSheet("font-size: 11px; color: #8888aa;")
        layout.addWidget(self.lbl_detect_status)

        # API Key
        layout.addWidget(QLabel("JX3API Token:"))
        token_box = QHBoxLayout()
        token_box.setSpacing(8)
        self.token_input = QLineEdit(self.config.get("api_key", ""))
        self.token_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.btn_toggle_token = QPushButton("👁 显示")
        self.btn_toggle_token.setFixedWidth(75)
        self.btn_toggle_token.setToolTip("切换密文/明文显示")
        self.btn_toggle_token.clicked.connect(self.toggle_token_visibility)
        token_box.addWidget(self.token_input)
        token_box.addWidget(self.btn_toggle_token)
        layout.addLayout(token_box)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_save = QPushButton("保存并进入")
        btn_save.setObjectName("PrimaryBtn")
        btn_save.clicked.connect(self.save)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

        # 初始检测与状态提示
        curr_p = self.path_input.text().strip()
        if is_valid_game_path(curr_p):
            self.set_status_success("✓ 当前游戏路径有效")
        else:
            self.run_detection()

    def set_status_success(self, text):
        self.lbl_detect_status.setText(text)
        self.lbl_detect_status.setStyleSheet("font-size: 11px; color: #44cc44;")

    def set_status_fail(self, text):
        self.lbl_detect_status.setText(text)
        self.lbl_detect_status.setStyleSheet("font-size: 11px; color: #8888aa;")

    def on_path_changed(self, text):
        p = text.strip()
        if is_valid_game_path(p):
            self.set_status_success("✓ 当前游戏路径有效（含 my#data）")
        elif p and os.path.exists(p):
            derived, _ = detect_game_path(p)
            if derived:
                self.lbl_detect_status.setText("提示：可推导至有效路径")
                self.lbl_detect_status.setStyleSheet("font-size: 11px; color: #ffaa00;")
            else:
                self.set_status_fail("未自动检测到，请手动浏览选择")
        elif not p:
            self.set_status_fail("未自动检测到，请手动浏览选择")
        else:
            self.set_status_fail("未自动检测到，请手动浏览选择")

    def run_detection(self):
        curr = self.path_input.text().strip()
        detected, source = detect_game_path(curr if curr else None)
        if detected:
            self.path_input.setText(detected)
            self.set_status_success(f"✓ 已自动检测到游戏路径（来源：{source}）")
        else:
            self.set_status_fail("未自动检测到，请手动浏览选择")

    def toggle_token_visibility(self):
        if self.token_input.echoMode() == QLineEdit.EchoMode.Password:
            self.token_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.btn_toggle_token.setText("🙈 隐藏")
        else:
            self.token_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.btn_toggle_token.setText("👁 显示")

    def browse_path(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择剑网3安装目录")
        if dir_path:
            if not is_valid_game_path(dir_path):
                derived, _ = detect_game_path(dir_path)
                if derived:
                    dir_path = derived
            self.path_input.setText(dir_path)

    def save(self):
        gpath = self.path_input.text().strip()
        key = self.token_input.text().strip()
        if not gpath or not os.path.exists(gpath):
            QMessageBox.warning(self, "提示", "请选择有效的剑网3游戏目录！")
            return
        if not is_valid_game_path(gpath):
            derived, _ = detect_game_path(gpath)
            if derived:
                gpath = derived
                self.path_input.setText(gpath)
        self.config["game_path"] = gpath
        self.config["api_key"] = key
        save_config(self.config)
        self.accept()


class ApiConfigDialog(QDialog):
    """查看或修改 JX3API Token 的配置对话框"""
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.token_modified = False
        self.setWindowTitle("API 设置")
        self.resize(480, 230)
        self.setStyleSheet(DARK_QSS)
        icon = get_app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(18, 18, 18, 18)

        title = QLabel("🔑 JX3API Token 设置")
        title.setStyleSheet("font-size: 15px; font-weight: bold; color: #3b8ed0;")
        layout.addWidget(title)

        # 提示与获取链接
        desc_lbl = QLabel("用于在线获取百战周排班、精耐招式及活动日历。")
        desc_lbl.setStyleSheet("color: #b0b0cc; font-size: 12px;")
        layout.addWidget(desc_lbl)

        link_lbl = QLabel('Token 获取地址：<a href="https://www.jx3api.com" style="color: #64b5f6; text-decoration: underline;">https://www.jx3api.com</a>')
        link_lbl.setOpenExternalLinks(True)
        link_lbl.setStyleSheet("font-size: 12px;")
        layout.addWidget(link_lbl)

        # Token 输入与显隐切换
        token_header = QLabel("JX3API Token:")
        token_header.setStyleSheet("color: #e0e0e0; font-size: 12px; font-weight: bold;")
        layout.addWidget(token_header)

        token_box = QHBoxLayout()
        token_box.setSpacing(8)
        self.token_input = QLineEdit(self.config.get("api_key", ""))
        self.token_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.token_input.setPlaceholderText("请输入 JX3API Token (如 jx3api::...)")

        self.btn_toggle_token = QPushButton("👁 显示")
        self.btn_toggle_token.setFixedWidth(75)
        self.btn_toggle_token.setToolTip("切换密文/明文显示")
        self.btn_toggle_token.clicked.connect(self.toggle_token_visibility)

        token_box.addWidget(self.token_input)
        token_box.addWidget(self.btn_toggle_token)
        layout.addLayout(token_box)

        layout.addStretch()

        # 操作按钮
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        btn_layout.addStretch()

        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)

        btn_save = QPushButton("保存")
        btn_save.setObjectName("PrimaryBtn")
        btn_save.clicked.connect(self.save)
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

    def toggle_token_visibility(self):
        if self.token_input.echoMode() == QLineEdit.EchoMode.Password:
            self.token_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.btn_toggle_token.setText("🙈 隐藏")
        else:
            self.token_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.btn_toggle_token.setText("👁 显示")

    def save(self):
        new_token = self.token_input.text().strip()
        old_token = (self.config.get("api_key") or "").strip()

        if new_token == old_token:
            QMessageBox.information(self, "提示", "Token 未修改")
            self.accept()
            return

        confirm = QMessageBox.question(
            self,
            "确认修改",
            "修改后立即生效，会影响百战排班/日历/精耐等在线功能，确认修改？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        # 更新配置
        self.config["api_key"] = new_token
        save_config(self.config)

        # 运行时同步
        try:
            from readers.baizhan_api import api as bz_api
            bz_api.api_key = new_token
        except Exception as e:
            logger.warning(f"同步运行时 API Token 失败: {e}")

        self.token_modified = True
        self.accept()


class DataLoaderThread(QThread):
    """后台数据加载线程，避免主 UI 卡死"""
    loaded = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, mgr):
        super().__init__()
        self.mgr = mgr

    def run(self):
        try:
            chars = self.mgr.load_all()
            self.loaded.emit(chars)
        except Exception as e:
            logger.error(f"Data loading failed: {e}")
            self.error.emit(str(e))


class ApiFetchThread(QThread):
    """后台 API 查询线程 (百战精耐/技能)"""
    fetched = pyqtSignal(str, dict)

    def __init__(self, mgr, name):
        super().__init__()
        self.mgr = mgr
        self.name = name

    def run(self):
        try:
            res = self.mgr.fetch_baizhan_info(self.name)
            self.fetched.emit(self.name, res)
        except Exception as e:
            logger.error(f"API fetch failed for {self.name}: {e}")
            self.fetched.emit(self.name, {"error": str(e)})

class CalendarFetchThread(QThread):
    """后台 API 查询线程 (周常活动日历)"""
    fetched = pyqtSignal(dict)

    def __init__(self, mgr):
        super().__init__()
        self.mgr = mgr

    def run(self):
        cal = self.mgr.fetch_active_calendar(force_refresh=True)
        self.fetched.emit(cal)


class RosterFetchThread(QThread):
    """后台 API 查询线程 (百战首领排班)"""
    fetched = pyqtSignal(dict)

    def __init__(self, mgr):
        super().__init__()
        self.mgr = mgr

    def run(self):
        wb = self.mgr.fetch_weekly_bosses(force_refresh=True)
        self.fetched.emit(wb)


class RoleDetailDialog(QDialog):
    """角色深度详情对话框（属性数据 / 角色背包 / 角色成就 / 角色宠物 / 角色奇遇）"""
    def __init__(self, char_data, data_path, parent=None):
        super().__init__(parent)
        self.char_data = char_data
        self.data_path = data_path
        from readers.detail_reader import RoleDetailReader
        self.detail_reader = RoleDetailReader(data_path)
        
        name = char_data.get("name", "角色详情")
        server = char_data.get("server", "")
        force_name = char_data.get("force_name", "")
        level = char_data.get("level", 120)
        
        self.setWindowTitle(f"⚔️ 【{name}】 角色全维度数据明细")
        self.resize(900, 640)
        self.setStyleSheet("""
            QDialog { background-color: #1e1e2d; color: #ffffff; font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; }
            QTabWidget::pane { border: 1px solid #333345; background: #252538; }
            QTabBar::tab { background: #181824; color: #aaa; padding: 10px 20px; font-weight: bold; border-top-left-radius: 6px; border-top-right-radius: 6px; }
            QTabBar::tab:selected { background: #3b8ed0; color: white; }
            QHeaderView::section { background-color: #161622; color: #ddd; font-weight: bold; padding: 6px; border: 1px solid #2a2a3c; }
            QTableWidget { background-color: #1e1e2d; color: #eee; gridline-color: #2e2e42; }
            QLineEdit { background-color: #161622; color: #fff; border: 1px solid #444460; padding: 6px 10px; border-radius: 4px; }
            QLabel { color: #eee; }
            QGroupBox { border: 1px solid #3d3d5c; border-radius: 6px; margin-top: 10px; font-weight: bold; color: #3b8ed0; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)
        
        # Banner Header
        banner = QFrame()
        banner.setStyleSheet("background-color: #28283d; border-radius: 6px; padding: 8px 12px;")
        b_layout = QHBoxLayout(banner)
        b_layout.setContentsMargins(0, 0, 0, 0)
        
        lbl_info = QLabel(f"👤 <span style='font-size: 15px;'><b>{name}</b></span>  ({char_data.get('region','')}/{server})  |  门派: <font color='#ffd54f'><b>{force_name}</b></font>  |  等级: <b>{level} 级</b>  |  装分: <font color='#4caf50'><b>{char_data.get('equip_score', 0):,}</b></font>")
        b_layout.addWidget(lbl_info)
        b_layout.addStretch()
        lbl_time = QLabel(f"📅 本地数据时刻: {char_data.get('last_update', '未知')}")
        lbl_time.setStyleSheet("color: #aaa; font-size: 12px;")
        b_layout.addWidget(lbl_time)
        layout.addWidget(banner)
        
        # Tabs
        self.tabs = QTabWidget()
        self.tabs.addTab(self.create_stats_tab(), "📊 角色属性数据")
        self.tabs.addTab(self.create_bag_tab(), "🎒 角色背包")
        self.tabs.addTab(self.create_achievements_tab(), "🏆 角色成就")
        self.tabs.addTab(self.create_pets_tab(), "🐾 角色宠物")
        self.tabs.addTab(self.create_serendipity_tab(), "🔮 角色奇遇")
        layout.addWidget(self.tabs)

    def create_stats_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        c = self.char_data
        name = c.get("name", "")
        server = c.get("server", "")

        gear_list, attrs = self.detail_reader.get_equipped_items(name, server)

        # Top Section: Combat & Basic Attributes Cards
        top_layout = QHBoxLayout()

        # Left Group Box: Basic & Resource Info
        gb_basic = QGroupBox("角色基础信息")
        grid_basic = QGridLayout()
        grid_basic.setSpacing(8)
        
        basic_items = [
            ("角色名称", c.get("name", "-")),
            ("区服/大区", f"{c.get('region', '')} / {c.get('server', '')}"),
            ("门派/等级", f"{c.get('force_name', '-')} ({c.get('level', 120)}级)"),
            ("成就资历点", f"{c.get('achievement_score', 0):,} 点"),
            ("持有金币", f"{c.get('gold', 0):,} 金"),
            ("休闲点(帮贡)", f"{c.get('contribution', 0):,} 点"),
            ("侠义值/威望", f"{c.get('justice', 0):,} 点"),
            ("数据更新时刻", c.get("last_update", "未知"))
        ]
        for idx, (label, val) in enumerate(basic_items):
            r, col = idx // 2, (idx % 2) * 2
            lbl_k = QLabel(f"{label}:")
            lbl_k.setStyleSheet("color: #aaa; font-size: 12px;")
            lbl_v = QLabel(str(val))
            lbl_v.setStyleSheet("color: #fff; font-size: 12px; font-weight: bold;")
            grid_basic.addWidget(lbl_k, r, col)
            grid_basic.addWidget(lbl_v, r, col + 1)
        gb_basic.setLayout(grid_basic)
        top_layout.addWidget(gb_basic)

        # Right Group Box: Combat Panel Stats
        gb_combat = QGroupBox("面板核心战斗属性")
        grid_combat = QGridLayout()
        grid_combat.setSpacing(8)

        combat_items = [
            ("总装备分数", f"<font color='#4caf50'><b>{c.get('equip_score', 0):,} 点</b></font>"),
            ("面板攻击力", f"<font color='#ffb74d'><b>{attrs.get('attack', 0):,}</b></font>"),
            ("会心等级", f"<font color='#ab47bc'><b>{attrs.get('crit', 0):,}</b></font>"),
            ("破防等级", f"<font color='#42a5f5'><b>{attrs.get('overcome', 0):,}</b></font>"),
            ("加速等级", f"<font color='#26a69a'><b>{attrs.get('haste', 0):,}</b></font>"),
            ("无双等级", f"<font color='#ef5350'><b>{attrs.get('strain', 0):,}</b></font>")
        ]
        for idx, (label, val) in enumerate(combat_items):
            r, col = idx // 2, (idx % 2) * 2
            lbl_k = QLabel(f"{label}:")
            lbl_k.setStyleSheet("color: #aaa; font-size: 12px;")
            lbl_v = QLabel(str(val))
            lbl_v.setStyleSheet("font-size: 12px;")
            grid_combat.addWidget(lbl_k, r, col)
            grid_combat.addWidget(lbl_v, r, col + 1)
        gb_combat.setLayout(grid_combat)
        top_layout.addWidget(gb_combat)

        layout.addLayout(top_layout)

        # Bottom Section: Equipped Items Table (12 slots)
        gb_gear = QGroupBox("已穿戴武器装备 (身上 12 个部位)")
        g_layout = QVBoxLayout()
        table_gear = QTableWidget()
        headers = ["部位", "装备名称", "精炼重数", "品质等级", "装备分数"]
        table_gear.setColumnCount(len(headers))
        table_gear.setHorizontalHeaderLabels(headers)
        table_gear.setAlternatingRowColors(True)
        table_gear.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table_gear.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        table_gear.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        table_gear.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        table_gear.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        main_gears = [g for g in gear_list if g["slot_id"] <= 12]
        table_gear.setRowCount(len(main_gears))
        for i, g in enumerate(main_gears):
            it_slot = QTableWidgetItem(g["slot_name"])
            it_name = QTableWidgetItem(g["name"])
            it_st = QTableWidgetItem(f"精炼 {g['strength']} 重" if g['strength'] > 0 else "未精炼")
            if g['strength'] > 0: it_st.setForeground(QBrush(QColor("#ffcc00")))
            
            it_q = NumericTableWidgetItem(str(g["quality_level"]), g["quality_level"])
            it_es = NumericTableWidgetItem(f"{g['equip_score']:,}", g["equip_score"])

            it_slot.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_st.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_q.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            it_es.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            table_gear.setItem(i, 0, it_slot)
            table_gear.setItem(i, 1, it_name)
            table_gear.setItem(i, 2, it_st)
            table_gear.setItem(i, 3, it_q)
            table_gear.setItem(i, 4, it_es)

        g_layout.addWidget(table_gear)
        gb_gear.setLayout(g_layout)
        layout.addWidget(gb_gear)

        return widget

    def create_bag_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # Search Bar
        search_bar = QHBoxLayout()
        search_bar.addWidget(QLabel("搜索背包物品:"))
        self.input_bag_search = QLineEdit()
        self.input_bag_search.setPlaceholderText("输入物品名称或描述筛选...")
        self.input_bag_search.textChanged.connect(self.apply_bag_filter)
        search_bar.addWidget(self.input_bag_search)
        layout.addLayout(search_bar)

        # Bag Items Table
        self.table_bag = QTableWidget()
        headers = ["物品名称", "数量", "品质", "详细说明"]
        self.table_bag.setColumnCount(len(headers))
        self.table_bag.setHorizontalHeaderLabels(headers)
        self.table_bag.setAlternatingRowColors(True)
        self.table_bag.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_bag.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table_bag.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table_bag.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table_bag)

        self.all_bag_items = self.detail_reader.get_bag_items(self.char_data.get("name", ""), self.char_data.get("server", ""))
        self.apply_bag_filter()
        return widget

    def apply_bag_filter(self):
        kw = self.input_bag_search.text().strip().lower()
        filtered = [x for x in self.all_bag_items if not kw or kw in x["name"].lower() or kw in x["desc"].lower()]
        
        self.table_bag.setSortingEnabled(False)
        self.table_bag.setRowCount(len(filtered))
        for i, item in enumerate(filtered):
            it_name = QTableWidgetItem(item["name"])
            it_cnt = NumericTableWidgetItem(str(item["count"]), item["count"])
            
            # Quality coloring
            q_color = "#ffffff"
            if item["quality"] == 5: q_color = "#ff8c00" # Orange
            elif item["quality"] == 4: q_color = "#a335ee" # Purple
            elif item["quality"] == 3: q_color = "#0070dd" # Blue
            elif item["quality"] == 2: q_color = "#1ef01e" # Green
            
            it_q = QTableWidgetItem(f"品质 {item['quality']}")
            it_q.setForeground(QBrush(QColor(q_color)))
            it_desc = QTableWidgetItem(item["desc"])
            
            self.table_bag.setItem(i, 0, it_name)
            self.table_bag.setItem(i, 1, it_cnt)
            self.table_bag.setItem(i, 2, it_q)
            self.table_bag.setItem(i, 3, it_desc)
        self.table_bag.setSortingEnabled(True)

    def create_achievements_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        uid = self.char_data.get("guid", "") or self.char_data.get("uid", "")
        achs = self.detail_reader.get_achievements(uid)

        lbl = QLabel(f"🏆 角色资历点数: <b><font color='#ffd54f'>{self.char_data.get('achievement_score', 0):,} 点</font></b>  |  已解锁成就快照数: <b>{len(achs)} 项</b>")
        lbl.setStyleSheet("font-size: 14px; padding: 6px; background: #28283d; border-radius: 4px;")
        layout.addWidget(lbl)

        table_ach = QTableWidget()
        table_ach.setColumnCount(2)
        table_ach.setHorizontalHeaderLabels(["成就 ID", "获取状态"])
        table_ach.setRowCount(len(achs))
        table_ach.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        for i, aid in enumerate(achs):
            it_id = QTableWidgetItem(f"成就 #{aid}")
            it_st = QTableWidgetItem("✓ 已达成")
            it_st.setForeground(QBrush(QColor("#4caf50")))
            table_ach.setItem(i, 0, it_id)
            table_ach.setItem(i, 1, it_st)

        layout.addWidget(table_ach)
        return widget

    def create_pets_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(15, 15, 15, 15)

        lbl = QLabel(f"🐾 角色宠物资历分数: <b><font color='#48bfe3'>{self.char_data.get('pet_score', 0):,} 点</font></b>")
        lbl.setStyleSheet("font-size: 14px; padding: 10px; background: #28283d; border-radius: 4px;")
        layout.addWidget(lbl)
        
        info = QLabel("宠物资历总积分正常解析展示。可在宠物界面联动显示详细捕获/拥有的宠物卡片。")
        info.setStyleSheet("color: #aaa; margin-top: 10px;")
        layout.addWidget(info)
        layout.addStretch()
        return widget

    def create_serendipity_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        uid = self.char_data.get("guid", "") or self.char_data.get("uid", "")
        records = self.detail_reader.get_serendipity_records(uid)

        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["奇遇名称", "触发时间", "状态"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setRowCount(len(records))

        for i, rec in enumerate(records):
            table.setItem(i, 0, QTableWidgetItem(rec["name"]))
            table.setItem(i, 1, QTableWidgetItem(rec["time"]))
            it_st = QTableWidgetItem("已完成")
            it_st.setForeground(QBrush(QColor("#ffd54f")))
            table.setItem(i, 2, it_st)

        layout.addWidget(table)
        return widget


class RosterNodeWidget(QFrame):
    """百战异闻录 蛇形关卡图谱单节点卡片（100层图谱图片同款设计）"""
    def __init__(self, floor_data, parent=None):
        super().__init__(parent)
        self.setFixedSize(76, 108)
        self.setStyleSheet("""
            RosterNodeWidget {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
            }
            RosterNodeWidget:hover {
                background-color: #f1f5f9;
                border: 1.5px solid #3b8ed0;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(2, 4, 2, 4)
        layout.setSpacing(1)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        idx = floor_data.get("index", 0)
        b_name = floor_data.get("name", "")
        d_info = floor_data.get("data", {}) or {}
        eff_name = d_info.get("name", "无")
        eff_desc = d_info.get("desc", "")

        tag_text = ""
        tag_color = "#64748b"
        orb_color = "#2563eb"

        EFFECT_TAGS = {
            "因陀罗的护佑": ("随机前进+100", "#16a34a", "#16a34a"),
            "阿修罗的悟性": ("秒杀+100", "#dc2626", "#dc2626"),
            "阿修罗的愤怒": ("后三+120", "#dc2626", "#e11d48"),
            "阿修罗的幸运": ("稀有提高+120", "#d97706", "#d97706"),
            "阿修罗的本性": ("+200", "#dc2626", "#dc2626"),
            "因陀罗的胜机": ("+500", "#d97706", "#d97706"),
            "因陀罗的策略": ("后六翻倍+50", "#c026d3", "#c026d3"),
            "因陀罗的战术": ("前六减半+50", "#0284c7", "#0284c7"),
            "因陀罗的迂回": ("逆向前进", "#16a34a", "#16a34a"),
        }

        if eff_name in EFFECT_TAGS:
            tag_text, tag_color, orb_color = EFFECT_TAGS[eff_name]
        elif idx <= 50:
            tag_text = "免消耗"
            tag_color = "#94a3b8"
            orb_color = "#2563eb"

        # Boss Avatar Icon (Replacing plain circle dot with real JX3Box Boss Avatar)
        lbl_orb = QLabel()
        lbl_orb.setFixedSize(36, 36)
        lbl_orb.setAlignment(Qt.AlignmentFlag.AlignCenter)

        NAME_MAP = {
            "上杉勇刀": "上衫勇刀",
            "恶凤灵霄峡": "恶战灵霄峡",
            "谢云流·青年": "谢云流",
            "程沐华·青年": "程沐华",
            "韦柔丝·困境": "韦柔丝",
        }
        target_name = NAME_MAP.get(b_name, b_name)

        base_dir = os.path.join(os.path.dirname(__file__), "data", "boss_avatars")
        avatar_path = os.path.join(base_dir, f"{target_name}.png")
        if not os.path.exists(avatar_path) or os.path.getsize(avatar_path) < 1000:
            avatar_path = os.path.join(base_dir, f"{b_name}.png")

        if not os.path.exists(avatar_path) or os.path.getsize(avatar_path) < 1000:
            try:
                from readers.boss_avatar import download_boss_avatar
                avatar_path = download_boss_avatar(b_name)
            except Exception:
                avatar_path = None

        if avatar_path and os.path.exists(avatar_path):
            pix = QPixmap(avatar_path)
            if not pix.isNull():
                scaled_pix = pix.scaled(36, 36, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                lbl_orb.setPixmap(scaled_pix)
                lbl_orb.setStyleSheet(f"border-radius: 18px; border: 2px solid {orb_color}; background-color: #0f172a;")
            else:
                lbl_orb.setStyleSheet(f"""
                    QLabel {{
                        background: qradialgradient(cx:0.3, cy:0.3, radius:0.85, fx:0.25, fy:0.25, stop:0 #ffffff, stop:0.45 {orb_color}, stop:1 #0f172a);
                        border-radius: 18px;
                    }}
                """)
        else:
            lbl_orb.setStyleSheet(f"""
                QLabel {{
                    background: qradialgradient(cx:0.3, cy:0.3, radius:0.85, fx:0.25, fy:0.25, stop:0 #ffffff, stop:0.45 {orb_color}, stop:1 #0f172a);
                    border-radius: 18px;
                }}
            """)

        layout.addWidget(lbl_orb, alignment=Qt.AlignmentFlag.AlignCenter)

        # Floor Index
        lbl_num = QLabel(f"{idx:02d}")
        lbl_num.setStyleSheet("font-size: 14px; font-weight: 800; color: #1e293b; border: none;")
        lbl_num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_num)

        # Boss Name
        lbl_boss = QLabel(b_name)
        lbl_boss.setStyleSheet("font-size: 11px; font-weight: bold; color: #334155; border: none;")
        lbl_boss.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_boss)

        # Effect Tag
        lbl_tag = QLabel(tag_text if tag_text else " ")
        lbl_tag.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {tag_color}; border: none;")
        lbl_tag.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_tag)

        # Tooltip
        skills = floor_data.get("skill", [])
        sk_str = ", ".join(skills) if skills else "无"
        tip = f"<b>第 {idx} 层 - {b_name}</b><br/>"
        if sk_str: tip += f"<b>主要招式:</b> {sk_str}<br/>"
        if eff_name != "无": tip += f"<b>特殊效果:</b> {eff_name}<br/><b>效果说明:</b> {eff_desc}"
        self.setToolTip(tip)


class StickyNoteDialog(QDialog):
    """便签式备注编辑弹窗"""
    def __init__(self, char_name, note_type_str, initial_text="", parent=None):
        super().__init__(parent)
        self.char_name = char_name
        self.note_type_str = note_type_str
        self.original_text = (initial_text or "").strip()
        self.result_text = self.original_text
        self._saved_flag = False

        self.setWindowTitle(f"📌 【{char_name}】 - {note_type_str}便签")
        self.resize(440, 300)
        self.setStyleSheet("""
            QDialog {
                background-color: #262419;
                border: 2px solid #6e5e2e;
                border-radius: 8px;
            }
            QLabel {
                color: #ffe082;
                font-family: 'Microsoft YaHei', sans-serif;
            }
            QTextEdit {
                background-color: #1a1910;
                color: #fff8e1;
                border: 1px solid #5c4e26;
                border-radius: 6px;
                padding: 10px;
                font-size: 13px;
                font-family: 'Microsoft YaHei', sans-serif;
                selection-background-color: #f57f17;
            }
            QTextEdit:focus {
                border: 1px solid #ffd54f;
            }
            QPushButton#SaveBtn {
                background-color: #f57f17;
                color: #ffffff;
                font-weight: bold;
                border-radius: 4px;
                padding: 6px 18px;
            }
            QPushButton#SaveBtn:hover {
                background-color: #ff8f00;
            }
            QPushButton#CancelBtn {
                background-color: #3e3a2a;
                color: #cccccc;
                border-radius: 4px;
                padding: 6px 14px;
            }
            QPushButton#CancelBtn:hover {
                background-color: #4e4836;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        # Header
        h_layout = QHBoxLayout()
        lbl_icon = QLabel("📌")
        lbl_icon.setStyleSheet("font-size: 18px;")
        h_layout.addWidget(lbl_icon)

        lbl_title = QLabel(f"<b>【{char_name}】</b> 的 <b>{note_type_str}便签</b>")
        lbl_title.setStyleSheet("font-size: 14px; color: #ffe082;")
        h_layout.addWidget(lbl_title)
        h_layout.addStretch()
        layout.addLayout(h_layout)

        # Text Edit
        self.text_edit = QTextEdit()
        self.text_edit.setPlainText(self.original_text)
        layout.addWidget(self.text_edit)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self.btn_save = QPushButton("💾 保存便签")
        self.btn_save.setObjectName("SaveBtn")
        self.btn_save.clicked.connect(self.on_save_clicked)
        btn_layout.addWidget(self.btn_save)

        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.setObjectName("CancelBtn")
        self.btn_cancel.clicked.connect(self.close)
        btn_layout.addWidget(self.btn_cancel)

        layout.addLayout(btn_layout)

    def on_save_clicked(self):
        self.result_text = self.text_edit.toPlainText().strip()
        self._saved_flag = True
        self.accept()

    def closeEvent(self, event):
        if self._saved_flag:
            event.accept()
            return

        current_text = self.text_edit.toPlainText().strip()
        if current_text != self.original_text:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("提示")
            msg_box.setText(f"是否保存角色【{self.char_name}】的【{self.note_type_str}】修改？")
            msg_box.setIcon(QMessageBox.Icon.Question)
            
            btn_save = msg_box.addButton("保存", QMessageBox.ButtonRole.AcceptRole)
            btn_discard = msg_box.addButton("不保存", QMessageBox.ButtonRole.DestructiveRole)
            btn_cancel = msg_box.addButton("取消", QMessageBox.ButtonRole.RejectRole)
            
            msg_box.setDefaultButton(btn_save)
            msg_box.exec()

            clicked = msg_box.clickedButton()
            if clicked == btn_save:
                self.result_text = current_text
                self._saved_flag = True
                event.accept()
                self.accept()
            elif clicked == btn_discard:
                event.accept()
                self.reject()
            else:  # Cancel
                event.ignore()
        else:
            event.accept()
            self.reject()


class SkillCardWidget(QFrame):
    """百战招式卡片组件"""
    def __init__(self, bg_c, border_c, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {bg_c};
                border: none;
                border-radius: 6px;
            }}
            QFrame:hover {{
                border: 1px solid #ffd54f;
            }}
        """)


class CompactSkillsExportWidget(QWidget):
    """百战招式导出专用紧凑长图 Widget"""
    def __init__(self, char_name, skill_list, title_tag="全技能汇总", parent=None):
        super().__init__(parent)
        self.char_name = char_name
        self.skill_list = skill_list
        self.title_tag = title_tag
        self.init_ui()

    def init_ui(self):
        self.setFixedWidth(780)
        self.setStyleSheet("""
            QWidget {
                background-color: #161622;
                color: #ffffff;
                font-family: 'Microsoft YaHei', 'Segoe UI', sans-serif;
            }
        """)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(14, 14, 14, 14)
        main_layout.setSpacing(10)

        # 头部 Banner
        header = QFrame()
        header.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #1a237e, stop:1 #311b92);
                border-radius: 6px;
            }
        """)
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(12, 8, 12, 8)

        title_lbl = QLabel(f"⚔️ <b>【{self.char_name}】百战招式一览 ({self.title_tag})</b>")
        title_lbl.setStyleSheet("font-size: 15px; font-weight: bold; color: #ffffff; background: transparent;")
        h_layout.addWidget(title_lbl)

        h_layout.addStretch()

        from datetime import datetime
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        count_lbl = QLabel(f"共 {len(self.skill_list)} 门招式  |  {now_str}")
        count_lbl.setStyleSheet("font-size: 11px; color: #b388ff; background: transparent;")
        h_layout.addWidget(count_lbl)

        main_layout.addWidget(header)

        # 按技能重数从高到低分组 (10 重 -> 0 重)
        from collections import defaultdict
        grouped_by_lvl = defaultdict(list)
        for sk in self.skill_list:
            lv = sk.get("nLevel", 0)
            grouped_by_lvl[lv].append(sk)

        sorted_lvls = sorted(grouped_by_lvl.keys(), reverse=True)

        for lv in sorted_lvls:
            sks = grouped_by_lvl[lv]
            # 组内按颜色与名称排序
            sks.sort(key=lambda s: (COLOR_ORDER.get(s.get("nColor", 0), (99, ""))[0], s.get("szSkillName") or s.get("szName") or ""))

            gbox = QGroupBox()
            gbox.setStyleSheet("""
                QGroupBox {
                    background-color: #1e1e2d;
                    border: 1px solid #2e2e42;
                    border-radius: 6px;
                    margin-top: 4px;
                    padding-top: 6px;
                }
            """)
            g_layout = QVBoxLayout(gbox)
            g_layout.setContentsMargins(8, 6, 8, 8)
            g_layout.setSpacing(6)

            # 组标题
            g_header = QLabel(f"⭐ <b>{lv} 重招式</b>  <font color='#8888aa'>({len(sks)} 门)</font>")
            g_header.setStyleSheet("font-size: 13px; font-weight: bold; color: #ffd54f; background: transparent;")
            g_layout.addWidget(g_header)

            # 网格布局 (5 列)
            grid = QGridLayout()
            grid.setSpacing(6)
            grid.setContentsMargins(0, 0, 0, 0)

            COLS = 5
            for idx, sk in enumerate(sks):
                r = idx // COLS
                c = idx % COLS

                card = self.create_compact_card(sk)
                grid.addWidget(card, r, c)

            g_layout.addLayout(grid)
            main_layout.addWidget(gbox)

        # 页脚落款
        footer = QLabel("Generated by 剑3小助手 · 百战招式大局览")
        footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer.setStyleSheet("color: #666688; font-size: 10px; margin-top: 4px;")
        main_layout.addWidget(footer)

        self.adjustSize()

    def create_compact_card(self, sk):
        sname = sk.get("szSkillName") or sk.get("szName") or "未知技能"
        slvl = sk.get("nLevel", 0)
        scol = sk.get("nColor", 0)
        iid = sk.get("dwInSkillID", 0)

        bg_c, border_c, text_c, label_c = COLOR_STYLES.get(scol, ("#2a2a3c", "#444460", "#ffffff", "#ffffff"))

        card = QFrame()
        card.setFixedSize(142, 42)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {bg_c};
                border: 1px solid {border_c};
                border-radius: 5px;
            }}
        """)

        c_layout = QHBoxLayout(card)
        c_layout.setContentsMargins(4, 3, 6, 3)
        c_layout.setSpacing(6)

        # 图标
        lbl_icon = QLabel()
        lbl_icon.setFixedSize(34, 34)
        
        ld = os.path.join(os.path.dirname(__file__), "web", "icons")
        cache_dir = os.path.join(os.path.dirname(__file__), "data", "bz_cache", "icons")
        icon_fp = os.path.join(ld, f"{sname}.png")
        if not os.path.exists(icon_fp) and iid:
            icon_fp = os.path.join(cache_dir, f"{iid}.png")

        if os.path.exists(icon_fp):
            pix = QPixmap(icon_fp)
            if not pix.isNull():
                lbl_icon.setPixmap(pix.scaled(34, 34, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                lbl_icon.setStyleSheet(f"border: 1px solid {border_c}; border-radius: 4px; background: transparent;")
            else:
                lbl_icon.setText(sname[:1])
                lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
                lbl_icon.setStyleSheet(f"background-color: {label_c}; color: #000; font-weight: bold; border-radius: 4px; font-size: 12px;")
        else:
            lbl_icon.setText(sname[:1])
            lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_icon.setStyleSheet(f"background-color: {label_c}; color: #000; font-weight: bold; border-radius: 4px; font-size: 12px;")

        c_layout.addWidget(lbl_icon)

        # 文字信息：仅技能名称与重数
        t_layout = QVBoxLayout()
        t_layout.setContentsMargins(0, 0, 0, 0)
        t_layout.setSpacing(1)

        lbl_name = QLabel(sname)
        lbl_name.setStyleSheet(f"color: {text_c}; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        t_layout.addWidget(lbl_name)

        lbl_lvl = QLabel(f"{slvl} 重")
        lbl_lvl.setStyleSheet("color: #ffd54f; font-size: 11px; font-weight: bold; background: transparent; border: none;")
        t_layout.addWidget(lbl_lvl)

        c_layout.addLayout(t_layout)
        return card


class CoreSkillsConfigDialog(QDialog):
    """百战技能分类配置弹窗"""
    def __init__(self, config_path: Optional[str] = None, parent=None):
        super().__init__(parent)
        self.config_path = config_path
        self.saved = False

        from bz_core_skills import (
            load_core_skill_categories,
            save_core_skill_categories,
            derive_core_skill_categories,
            get_all_known_skills,
            get_skill_meta,
            CORE_CATEGORY_SLOTS,
        )
        self._load_core_skill_categories = load_core_skill_categories
        self._save_core_skill_categories = save_core_skill_categories
        self._derive_core_skill_categories = derive_core_skill_categories
        self._get_all_known_skills = get_all_known_skills
        self._get_skill_meta = get_skill_meta
        self._core_category_slots = CORE_CATEGORY_SLOTS

        # 深拷贝载入配置，避免编辑过程中直接污染
        raw_cats = self._load_core_skill_categories(self.config_path)
        self.categories = [
            {
                "group": str(c.get("group", "")),
                "window": str(c.get("window", "")),
                "candidates": list(c.get("candidates", [])),
                "enabled": bool(c.get("enabled", True)),
                "display_count": max(1, int(c.get("display_count", 1))),
            }
            for c in raw_cats
        ]

        self.all_skills = self._get_all_known_skills()
        self._skill_meta_cache = {}
        for sname in self.all_skills:
            self._skill_meta_cache[sname] = self._get_skill_meta(sname)

        # 招式调息时间（jx3box 数据，覆盖全部 156 个）与消耗点数（占用技能格），用于标注与筛选
        try:
            from bz_core_skills import get_skill_cooldowns, get_skill_costs, get_level_colors
            self._skill_cds = get_skill_cooldowns()
            self._skill_costs = get_skill_costs()
            self._level_colors = get_level_colors(self.config_path)
        except Exception:
            self._skill_cds = {}
            self._skill_costs = {}
            self._level_colors = []

        self.current_cat_idx = -1
        self._is_updating_ui = False

        self.setWindowTitle("⚙ 百战技能分类配置")
        self.resize(1020, 700)
        icon = get_app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)

        self.setStyleSheet("""
            QDialog { background-color: #1e1e2d; color: #ffffff; font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; }
            QLabel { color: #eee; }
            QPushButton#PrimaryBtn { background-color: #0d47a1; color: white; font-weight: bold; border-radius: 4px; padding: 6px 14px; }
            QPushButton#PrimaryBtn:hover { background-color: #1565c0; }
            QPushButton#GreenBtn { background-color: #2e7d32; color: white; font-weight: bold; border-radius: 4px; padding: 6px 14px; }
            QPushButton#GreenBtn:hover { background-color: #388e3c; }
            QPushButton#DangerBtn { background-color: #8e2424; color: white; font-weight: bold; border-radius: 4px; padding: 5px 10px; }
            QPushButton#DangerBtn:hover { background-color: #b71c1c; }
            QPushButton#DefaultBtn { background-color: #2b2b3d; color: #e0e0e0; border: 1px solid #444460; border-radius: 4px; padding: 6px 12px; }
            QPushButton#DefaultBtn:hover { background-color: #38384f; }
            QPushButton#SmallBtn { background-color: #2b2b3d; color: #ffffff; border: 1px solid #444460; border-radius: 4px; padding: 4px 8px; font-size: 12px; }
            QPushButton#SmallBtn:hover { background-color: #38384f; }
            QListWidget { background-color: #161622; color: #ffffff; border: 1px solid #333345; border-radius: 6px; }
            QListWidget::item { padding: 4px 6px; }
            QListWidget::item:selected { background-color: #2b3b5c; }
            QLineEdit, QSpinBox { background-color: #161622; color: #ffffff; border: 1px solid #444460; border-radius: 4px; padding: 4px 8px; }
            QScrollBar:vertical { background: #161622; width: 10px; }
            QScrollBar::handle:vertical { background: #3b3b54; border-radius: 5px; }
            QScrollBar:horizontal { background: #161622; height: 10px; }
            QScrollBar::handle:horizontal { background: #3b3b54; border-radius: 5px; }
        """)

        self.init_ui()

    def _format_skill_label(self, sname: str) -> str:
        cd = self._skill_cds.get(sname)
        cost = self._skill_costs.get(sname)
        if cd is None:
            cd_str = "-"
        elif cd == 0:
            cd_str = "无调息"
        else:
            cd_str = f"{cd}秒"
        cost_str = f" {cost}点" if cost is not None else ""
        return f"{sname}  ·{cd_str}{cost_str}"

    def _get_skill_tooltip(self, sname: str) -> str:
        meta = self._skill_meta_cache.get(sname) or self._get_skill_meta(sname)
        detail = meta.get("detail", "")
        cd = self._skill_cds.get(sname)
        cost = self._skill_costs.get(sname)
        lines = [f"【{sname}】"]
        if cd is None:
            lines.append("【调息时间】: -（无数据，默认归 10S 档，可手动划分）")
        elif cd == 0:
            lines.append("【调息时间】: 无调息时间")
        else:
            win = "10S" if cd <= 10 else ("30S" if cd <= 30 else "1分钟")
            lines.append(f"【调息时间】: {cd} 秒（对应 {win} 档）")
        if cost is not None:
            lines.append(f"【消耗点数】: {cost} 点（占用 {cost} 个技能格）")
        if detail:
            lines.append(detail[:200] + ("..." if len(detail) > 200 else ""))
        return "\n".join(lines)

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        content_layout = QHBoxLayout()
        content_layout.setSpacing(14)

        # ------------------- 左侧分类列表区 (~320px) -------------------
        left_box = QVBoxLayout()
        left_box.setSpacing(8)

        lbl_left_title = QLabel("📁 百战技能分类列表（勾选表示在总览展示）")
        lbl_left_title.setStyleSheet("font-weight: bold; font-size: 13px; color: #40a9ff;")
        left_box.addWidget(lbl_left_title)

        self.list_categories = QListWidget()
        self.list_categories.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.list_categories.currentRowChanged.connect(self.on_category_selected)
        self.list_categories.itemChanged.connect(self.on_category_item_changed)
        left_box.addWidget(self.list_categories)

        # 左侧操作按钮网格
        btn_grid_top = QHBoxLayout()
        btn_grid_top.setSpacing(6)
        self.btn_add_cat = QPushButton("➕ 新增分类")
        self.btn_add_cat.setObjectName("SmallBtn")
        self.btn_add_cat.clicked.connect(self.add_category)
        btn_grid_top.addWidget(self.btn_add_cat)

        self.btn_copy_cat = QPushButton("📋 复制")
        self.btn_copy_cat.setObjectName("SmallBtn")
        self.btn_copy_cat.clicked.connect(self.copy_category)
        btn_grid_top.addWidget(self.btn_copy_cat)

        self.btn_rename_cat = QPushButton("✏ 重命名")
        self.btn_rename_cat.setObjectName("SmallBtn")
        self.btn_rename_cat.clicked.connect(self.rename_category)
        btn_grid_top.addWidget(self.btn_rename_cat)

        self.btn_del_cat = QPushButton("🗑 删除")
        self.btn_del_cat.setObjectName("DangerBtn")
        self.btn_del_cat.clicked.connect(self.delete_category)
        btn_grid_top.addWidget(self.btn_del_cat)

        left_box.addLayout(btn_grid_top)

        btn_grid_order = QHBoxLayout()
        btn_grid_order.setSpacing(6)
        self.btn_cat_up = QPushButton("↑ 上移分类")
        self.btn_cat_up.setObjectName("SmallBtn")
        self.btn_cat_up.clicked.connect(self.move_up_category)
        btn_grid_order.addWidget(self.btn_cat_up)

        self.btn_cat_down = QPushButton("↓ 下移分类")
        self.btn_cat_down.setObjectName("SmallBtn")
        self.btn_cat_down.clicked.connect(self.move_down_category)
        btn_grid_order.addWidget(self.btn_cat_down)

        left_box.addLayout(btn_grid_order)

        left_widget = QWidget()
        left_widget.setLayout(left_box)
        left_widget.setFixedWidth(330)
        content_layout.addWidget(left_widget)

        # ------------------- 右侧编辑区 -------------------
        self.right_stack = QStackedWidget()

        # Page 0: 未选中任何分类时的占位
        empty_page = QWidget()
        empty_layout = QVBoxLayout(empty_page)
        lbl_empty = QLabel("👈 请在左侧选择或新增一个技能分类进行详细配置")
        lbl_empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_empty.setStyleSheet("font-size: 15px; color: #777799;")
        empty_layout.addWidget(lbl_empty)
        self.right_stack.addWidget(empty_page)

        # Page 1: 具体的分类编辑器
        edit_page = QWidget()
        edit_layout = QVBoxLayout(edit_page)
        edit_layout.setContentsMargins(4, 0, 0, 0)
        edit_layout.setSpacing(10)

        # 顶部：当前分类标题 + 展示技能数量
        top_edit_bar = QHBoxLayout()
        top_edit_bar.setSpacing(12)

        self.lbl_cat_title = QLabel("📌 分类：-")
        self.lbl_cat_title.setStyleSheet("font-size: 15px; font-weight: bold; color: #69c0ff;")
        top_edit_bar.addWidget(self.lbl_cat_title)

        top_edit_bar.addStretch()

        lbl_spin = QLabel("展示技能数量:")
        lbl_spin.setStyleSheet("font-weight: bold;")
        top_edit_bar.addWidget(lbl_spin)

        self.spin_display_count = QSpinBox()
        self.spin_display_count.setRange(1, 10)
        self.spin_display_count.setValue(5)
        self.spin_display_count.setFixedWidth(65)
        self.spin_display_count.setToolTip("该分类在总览表格中展示前 N 个已学最高等级技能（默认 5）")
        self.spin_display_count.valueChanged.connect(self.on_display_count_changed)
        top_edit_bar.addWidget(self.spin_display_count)

        # 等级颜色规则编辑（3 行：min/max/color），写入 bz_core_skills.json 的 level_colors
        lbl_lvc = QLabel("等级颜色:")
        lbl_lvc.setStyleSheet("font-weight: bold;")
        top_edit_bar.addWidget(lbl_lvc)

        self._lv_color_rows = []  # [(spin_min, spin_max, btn_color), ...]
        lv_colors_box = QHBoxLayout()
        lv_colors_box.setSpacing(4)
        for i in range(3):
            row = []
            smin = QSpinBox(); smin.setRange(1, 99); smin.setFixedWidth(46)
            smax = QSpinBox(); smax.setRange(1, 999); smax.setFixedWidth(46)
            btn = QPushButton(); btn.setFixedSize(30, 22)
            btn.setFlat(True)
            lv_colors_box.addWidget(QLabel("Lv" if i == 0 else ""))
            lv_colors_box.addWidget(smin)
            lv_colors_box.addWidget(QLabel("-" if i == 0 else "~"))
            lv_colors_box.addWidget(smax)
            lv_colors_box.addWidget(btn)
            row = (smin, smax, btn)
            self._lv_color_rows.append(row)
            btn.clicked.connect(lambda _=False, idx=i: self._pick_lv_color(idx))
        top_edit_bar.addLayout(lv_colors_box)

        edit_layout.addLayout(top_edit_bar)

        # 中部：双列表候选编辑器
        dual_lists_layout = QHBoxLayout()
        dual_lists_layout.setSpacing(10)

        # 左列表：全部技能候选池
        all_skills_box = QVBoxLayout()
        all_skills_box.setSpacing(6)

        lbl_all_pool = QLabel(f"全部已知技能 ({len(self.all_skills)} 个)")
        lbl_all_pool.setStyleSheet("font-weight: bold; color: #b0b8c8;")
        all_skills_box.addWidget(lbl_all_pool)

        self.txt_skill_search = QLineEdit()
        self.txt_skill_search.setPlaceholderText("🔍 搜索技能名称...")
        self.txt_skill_search.textChanged.connect(self.filter_all_skills)
        all_skills_box.addWidget(self.txt_skill_search)

        # 消耗点数筛选器
        cost_filter_layout = QHBoxLayout()
        cost_filter_layout.setSpacing(6)
        lbl_cost = QLabel("消耗点数:")
        lbl_cost.setStyleSheet("color: #b0b8c8; font-size: 12px;")
        self.cbo_cost_filter = QComboBox()
        self.cbo_cost_filter.addItems(["全部", "1点", "2点", "3点"])
        self.cbo_cost_filter.currentTextChanged.connect(self.filter_all_skills)
        cost_filter_layout.addWidget(lbl_cost)
        cost_filter_layout.addWidget(self.cbo_cost_filter)
        cost_filter_layout.addStretch()
        all_skills_box.addLayout(cost_filter_layout)

        self.list_all_skills = QListWidget()
        self.list_all_skills.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.list_all_skills.itemDoubleClicked.connect(self.on_all_skill_double_clicked)
        all_skills_box.addWidget(self.list_all_skills)

        dual_lists_layout.addLayout(all_skills_box, 1)

        # 中间操作按钮
        mid_btns_box = QVBoxLayout()
        mid_btns_box.setSpacing(10)
        mid_btns_box.addStretch()

        self.btn_add_cand = QPushButton("添加 ➡")
        self.btn_add_cand.setObjectName("SmallBtn")
        self.btn_add_cand.setToolTip("将左侧选中的技能加入本分类候选列表")
        self.btn_add_cand.clicked.connect(self.add_selected_candidates)
        mid_btns_box.addWidget(self.btn_add_cand)

        self.btn_remove_cand = QPushButton("⬅ 移除")
        self.btn_remove_cand.setObjectName("SmallBtn")
        self.btn_remove_cand.setToolTip("从右侧候选列表中移除选中的技能")
        self.btn_remove_cand.clicked.connect(self.remove_selected_candidates)
        mid_btns_box.addWidget(self.btn_remove_cand)

        mid_btns_box.addSpacing(16)

        self.btn_derive_slot = QPushButton("⚡ 自动推导")
        self.btn_derive_slot.setObjectName("DefaultBtn")
        self.btn_derive_slot.setToolTip("按规则推导当前 group·window 的标准候选技能")
        self.btn_derive_slot.clicked.connect(self.derive_current_slot)
        mid_btns_box.addWidget(self.btn_derive_slot)

        mid_btns_box.addStretch()
        dual_lists_layout.addLayout(mid_btns_box)

        # 右列表：本分类候选技能
        cand_box = QVBoxLayout()
        cand_box.setSpacing(6)

        self.lbl_cand_title = QLabel("本分类候选技能（按优先级排序）")
        self.lbl_cand_title.setStyleSheet("font-weight: bold; color: #52c41a;")
        cand_box.addWidget(self.lbl_cand_title)

        self.list_candidates = QListWidget()
        self.list_candidates.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.list_candidates.itemDoubleClicked.connect(self.on_candidate_double_clicked)
        cand_box.addWidget(self.list_candidates)

        cand_order_bar = QHBoxLayout()
        cand_order_bar.setSpacing(8)

        self.btn_cand_up = QPushButton("↑ 上移优先级")
        self.btn_cand_up.setObjectName("SmallBtn")
        self.btn_cand_up.setToolTip("提高候选技能优先级（等级相同时排在前面的优先展示）")
        self.btn_cand_up.clicked.connect(self.move_up_candidate)
        cand_order_bar.addWidget(self.btn_cand_up)

        self.btn_cand_down = QPushButton("↓ 下移优先级")
        self.btn_cand_down.setObjectName("SmallBtn")
        self.btn_cand_down.setToolTip("降低候选技能优先级")
        self.btn_cand_down.clicked.connect(self.move_down_candidate)
        cand_order_bar.addWidget(self.btn_cand_down)

        cand_box.addLayout(cand_order_bar)

        dual_lists_layout.addLayout(cand_box, 1)

        edit_layout.addLayout(dual_lists_layout)
        self.right_stack.addWidget(edit_page)

        content_layout.addWidget(self.right_stack, 1)
        main_layout.addLayout(content_layout)

        # ------------------- 底部按钮区 -------------------
        bottom_bar = QHBoxLayout()
        bottom_bar.setSpacing(12)

        self.btn_reset_default = QPushButton("🔄 恢复默认 7 档")
        self.btn_reset_default.setObjectName("DefaultBtn")
        self.btn_reset_default.setToolTip("使用自动推导规则恢复默认的 7 档分类配置")
        self.btn_reset_default.clicked.connect(self.reset_to_default)
        bottom_bar.addWidget(self.btn_reset_default)

        bottom_bar.addStretch()

        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.setObjectName("DefaultBtn")
        self.btn_cancel.clicked.connect(self.reject)
        bottom_bar.addWidget(self.btn_cancel)

        self.btn_save = QPushButton("💾 保存配置")
        self.btn_save.setObjectName("PrimaryBtn")
        self.btn_save.clicked.connect(self.save_config)
        bottom_bar.addWidget(self.btn_save)

        main_layout.addLayout(bottom_bar)

        # 底部数据说明
        lbl_data_note = QLabel(
            "ℹ️ 档位按 jx3box 调息时间划分：10秒 → 10S，30秒 → 30S，60秒 → 1分钟；"
            "25/50/300秒就近归档，无调息/无数据归 10S 档可手动划分。\n"
            "　　消耗点数 = 占用的技能格数（数据来自 jx3box，覆盖全部 156 个招式）。"
            "百战最多 3 个技能槽位，携带招式点数合计 ≤3。"
            "打击类型（打精/打耐/回复）取自招式描述。"
        )
        lbl_data_note.setStyleSheet("font-size: 11px; color: #8888aa; margin-top: 6px; line-height: 1.5;")
        lbl_data_note.setWordWrap(True)
        main_layout.addWidget(lbl_data_note)

        # 填充全部技能列表
        self.populate_all_skills_list()

        # 填充左侧分类列表
        self.refresh_categories_list(select_idx=0 if self.categories else -1)

    def populate_all_skills_list(self):
        self.list_all_skills.clear()
        for sname in self.all_skills:
            item = QListWidgetItem(self._format_skill_label(sname))
            item.setData(Qt.ItemDataRole.UserRole, sname)
            item.setToolTip(self._get_skill_tooltip(sname))
            self.list_all_skills.addItem(item)

    def filter_all_skills(self, _=None):
        """按搜索词 + 消耗点数联合筛选左侧技能池"""
        query = self.txt_skill_search.text().strip().lower()
        cost_sel = self.cbo_cost_filter.currentText() if hasattr(self, "cbo_cost_filter") else "全部"

        for i in range(self.list_all_skills.count()):
            item = self.list_all_skills.item(i)
            sname = item.data(Qt.ItemDataRole.UserRole) or item.text()

            hide = bool(query and query not in sname.lower())

            if not hide and cost_sel != "全部":
                cost = self._skill_costs.get(sname)
                try:
                    want = int(cost_sel.rstrip("点"))
                    hide = (cost != want)
                except ValueError:
                    pass

            item.setHidden(hide)

    def _cat_item_text(self, cat: dict) -> str:
        grp = cat.get("group", "")
        win = cat.get("window", "")
        c_count = len(cat.get("candidates", []))
        d_count = cat.get("display_count", 1)
        return f"{grp}·{win}  ({c_count}候选 / 显示{d_count}个)"

    def refresh_categories_list(self, select_idx: int = -1):
        self._is_updating_ui = True
        self.list_categories.clear()

        for cat in self.categories:
            item = QListWidgetItem(self._cat_item_text(cat))
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if cat.get("enabled", True) else Qt.CheckState.Unchecked)
            self.list_categories.addItem(item)

        self._is_updating_ui = False

        if 0 <= select_idx < len(self.categories):
            self.list_categories.setCurrentRow(select_idx)
        elif len(self.categories) > 0:
            self.list_categories.setCurrentRow(0)
        else:
            self.on_category_selected(-1)

    def update_current_cat_item_text(self):
        if 0 <= self.current_cat_idx < self.list_categories.count() and self.current_cat_idx < len(self.categories):
            self._is_updating_ui = True
            cat = self.categories[self.current_cat_idx]
            item = self.list_categories.item(self.current_cat_idx)
            item.setText(self._cat_item_text(cat))
            self._is_updating_ui = False

    def on_category_item_changed(self, item: QListWidgetItem):
        if self._is_updating_ui:
            return
        row = self.list_categories.row(item)
        if 0 <= row < len(self.categories):
            self.categories[row]["enabled"] = (item.checkState() == Qt.CheckState.Checked)

    def on_category_selected(self, row: int):
        self.current_cat_idx = row
        if row < 0 or row >= len(self.categories):
            self.right_stack.setCurrentIndex(0)
            return

        self.right_stack.setCurrentIndex(1)
        cat = self.categories[row]
        self.lbl_cat_title.setText(f"📌 分类：{cat.get('group', '')} · {cat.get('window', '')}")

        self._is_updating_ui = True
        self.spin_display_count.setValue(cat.get("display_count", 5))
        self._refresh_lv_color_rows()
        self._is_updating_ui = False

        self.refresh_candidates_list()

    def _refresh_lv_color_rows(self):
        """把 self._level_colors 规则刷到 3 行颜色编辑器上"""
        rules = (self._level_colors or [])[:3]
        while len(rules) < 3:
            rules.append({"min": 1, "max": 1, "color": "#888888"})
        for i, (smin, smax, btn) in enumerate(self._lv_color_rows):
            r = rules[i]
            smin.setValue(int(r["min"]))
            smax.setValue(int(r["max"]))
            btn.setStyleSheet(f"background-color: {r['color']}; border: 1px solid #666;")

    def _pick_lv_color(self, idx: int):
        """弹出取色器修改第 idx 行的颜色"""
        from PyQt6.QtWidgets import QColorDialog
        smin, smax, btn = self._lv_color_rows[idx]
        cur = QColor(btn.styleSheet().split("background-color:")[-1].split(";")[0].strip() or "#ffffff")
        color = QColorDialog.getColor(cur, self, "选择该等级段颜色")
        if color.isValid():
            btn.setStyleSheet(f"background-color: {color.name()}; border: 1px solid #666;")
            # 立即写回 self._level_colors
            rules = list(self._level_colors or [])[:3]
            while len(rules) < 3:
                rules.append({"min": 1, "max": 1, "color": "#888888"})
            rules[idx]["color"] = color.name()
            self._level_colors = rules

    def on_display_count_changed(self, val: int):
        if self._is_updating_ui or self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        self.categories[self.current_cat_idx]["display_count"] = val
        self.update_current_cat_item_text()

    def _collect_lv_colors(self):
        """从 3 行编辑器收集等级颜色规则，返回规则列表（自上而下匹配）"""
        rules = []
        for smin, smax, btn in self._lv_color_rows:
            mn, mx = smin.value(), smax.value()
            if mn > mx:
                mn, mx = mx, mn
            m = re.search(r"background-color:\s*(#[0-9a-fA-F]{3,8})", btn.styleSheet())
            color = m.group(1) if m else "#888888"
            rules.append({"min": mn, "max": mx, "color": color})
        return rules

    def refresh_candidates_list(self):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            self.list_candidates.clear()
            return

        cat = self.categories[self.current_cat_idx]
        candidates = cat.get("candidates", [])
        self.lbl_cand_title.setText(f"本分类候选技能 ({len(candidates)} 个，按优先级从高到低)")

        self.list_candidates.clear()
        for sname in candidates:
            item = QListWidgetItem(self._format_skill_label(sname))
            item.setData(Qt.ItemDataRole.UserRole, sname)
            item.setToolTip(self._get_skill_tooltip(sname))
            self.list_candidates.addItem(item)

    def add_category(self):
        grp, ok1 = QInputDialog.getText(self, "新增分类", "请输入分类名 (group)：\n例如：打精、打耐、回复、输出、控制")
        if not ok1 or not grp.strip():
            return
        win, ok2 = QInputDialog.getText(self, "新增分类", "请输入档位名 (window)：\n例如：1分钟、30S、10S、核心、爆发")
        if not ok2 or not win.strip():
            return

        grp = grp.strip()
        win = win.strip()
        for c in self.categories:
            if c.get("group") == grp and c.get("window") == win:
                QMessageBox.warning(self, "提示", f"已存在相同分类【{grp}·{win}】，不能重复添加。")
                return

        new_cat = {
            "group": grp,
            "window": win,
            "candidates": [],
            "enabled": True,
            "display_count": 1,
        }
        self.categories.append(new_cat)
        self.refresh_categories_list(select_idx=len(self.categories) - 1)

    def copy_category(self):
        idx = self.list_categories.currentRow()
        if idx < 0 or idx >= len(self.categories):
            return
        cur = self.categories[idx]

        win, ok = QInputDialog.getText(
            self, "复制分类",
            f"为分类【{cur.get('group', '')}】输入新档位名：",
            text=f"{cur.get('window', '')}_副本"
        )
        if not ok or not win.strip():
            return

        win = win.strip()
        grp = cur.get("group", "")
        for c in self.categories:
            if c.get("group") == grp and c.get("window") == win:
                QMessageBox.warning(self, "提示", f"已存在相同分类【{grp}·{win}】。")
                return

        new_cat = {
            "group": grp,
            "window": win,
            "candidates": list(cur.get("candidates", [])),
            "enabled": cur.get("enabled", True),
            "display_count": cur.get("display_count", 1),
        }
        self.categories.insert(idx + 1, new_cat)
        self.refresh_categories_list(select_idx=idx + 1)

    def rename_category(self):
        idx = self.list_categories.currentRow()
        if idx < 0 or idx >= len(self.categories):
            return
        cur = self.categories[idx]

        grp, ok1 = QInputDialog.getText(self, "重命名分类", "请输入分类名 (group)：", text=cur.get("group", ""))
        if not ok1 or not grp.strip():
            return
        win, ok2 = QInputDialog.getText(self, "重命名分类", "请输入档位名 (window)：", text=cur.get("window", ""))
        if not ok2 or not win.strip():
            return

        grp = grp.strip()
        win = win.strip()
        for i, c in enumerate(self.categories):
            if i != idx and c.get("group") == grp and c.get("window") == win:
                QMessageBox.warning(self, "提示", f"已存在相同分类【{grp}·{win}】。")
                return

        cur["group"] = grp
        cur["window"] = win
        self.refresh_categories_list(select_idx=idx)

    def delete_category(self):
        idx = self.list_categories.currentRow()
        if idx < 0 or idx >= len(self.categories):
            return
        cur = self.categories[idx]

        res = QMessageBox.question(
            self, "确认删除",
            f"确定要删除分类【{cur.get('group', '')}·{cur.get('window', '')}】吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if res == QMessageBox.StandardButton.Yes:
            del self.categories[idx]
            next_idx = min(idx, len(self.categories) - 1)
            self.refresh_categories_list(select_idx=next_idx)

    def move_up_category(self):
        idx = self.list_categories.currentRow()
        if idx > 0:
            self.categories[idx], self.categories[idx - 1] = self.categories[idx - 1], self.categories[idx]
            self.refresh_categories_list(select_idx=idx - 1)

    def move_down_category(self):
        idx = self.list_categories.currentRow()
        if 0 <= idx < len(self.categories) - 1:
            self.categories[idx], self.categories[idx + 1] = self.categories[idx + 1], self.categories[idx]
            self.refresh_categories_list(select_idx=idx + 1)

    def on_all_skill_double_clicked(self, item: QListWidgetItem):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        sname = item.data(Qt.ItemDataRole.UserRole) or item.text()
        cands = self.categories[self.current_cat_idx]["candidates"]
        if sname not in cands:
            cands.append(sname)
            self.refresh_candidates_list()
            self.update_current_cat_item_text()

    def on_candidate_double_clicked(self, item: QListWidgetItem):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        sname = item.data(Qt.ItemDataRole.UserRole) or item.text()
        cands = self.categories[self.current_cat_idx]["candidates"]
        if sname in cands:
            cands.remove(sname)
            self.refresh_candidates_list()
            self.update_current_cat_item_text()

    def add_selected_candidates(self):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        cands = self.categories[self.current_cat_idx]["candidates"]
        changed = False
        for item in self.list_all_skills.selectedItems():
            sname = item.data(Qt.ItemDataRole.UserRole) or item.text()
            if sname not in cands:
                cands.append(sname)
                changed = True
        if changed:
            self.refresh_candidates_list()
            self.update_current_cat_item_text()

    def remove_selected_candidates(self):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        cands = self.categories[self.current_cat_idx]["candidates"]
        changed = False
        for item in self.list_candidates.selectedItems():
            sname = item.data(Qt.ItemDataRole.UserRole) or item.text()
            if sname in cands:
                cands.remove(sname)
                changed = True
        if changed:
            self.refresh_candidates_list()
            self.update_current_cat_item_text()

    def move_up_candidate(self):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        idx = self.list_candidates.currentRow()
        cands = self.categories[self.current_cat_idx]["candidates"]
        if idx > 0 and idx < len(cands):
            cands[idx], cands[idx - 1] = cands[idx - 1], cands[idx]
            self.refresh_candidates_list()
            self.list_candidates.setCurrentRow(idx - 1)

    def move_down_candidate(self):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        idx = self.list_candidates.currentRow()
        cands = self.categories[self.current_cat_idx]["candidates"]
        if 0 <= idx < len(cands) - 1:
            cands[idx], cands[idx + 1] = cands[idx + 1], cands[idx]
            self.refresh_candidates_list()
            self.list_candidates.setCurrentRow(idx + 1)

    def derive_current_slot(self):
        if self.current_cat_idx < 0 or self.current_cat_idx >= len(self.categories):
            return
        cur = self.categories[self.current_cat_idx]
        grp = cur.get("group", "")
        win = cur.get("window", "")

        derived_all = self._derive_core_skill_categories()
        matched = next((d for d in derived_all if d.get("group") == grp and d.get("window") == win), None)

        if matched:
            cur["candidates"] = list(matched.get("candidates", []))
            self.refresh_candidates_list()
            self.update_current_cat_item_text()
            QMessageBox.information(self, "推导完成", f"已按规则为【{grp}·{win}】推导并载入 {len(cur['candidates'])} 个候选技能。")
        else:
            QMessageBox.information(self, "提示", "该自定义分类无自动推导规则")

    def reset_to_default(self):
        res = QMessageBox.question(
            self, "确认恢复默认",
            "确定要恢复为默认的 7 档分类配置吗？\n默认展示各档满级打击值最高的前 5 个技能，\n当前所有自定义分类与修改都将丢失。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if res == QMessageBox.StandardButton.Yes:
            self.categories = self._derive_core_skill_categories()
            self.refresh_categories_list(select_idx=0)

    def save_config(self):
        # 确保 enabled 状态与复选框完全同步
        for i in range(self.list_categories.count()):
            item = self.list_categories.item(i)
            if i < len(self.categories):
                self.categories[i]["enabled"] = (item.checkState() == Qt.CheckState.Checked)

        ok = self._save_core_skill_categories(
            self.categories,
            config_path=self.config_path,
            level_colors=self._collect_lv_colors(),
        )
        if ok:
            self.saved = True
            self.accept()
        else:
            QMessageBox.warning(self, "保存失败", "写入配置文件时发生错误，请检查文件权限。")


class AllAccountsBaizhanDialog(QDialog):
    """百战招式全账号总览弹窗"""
    def __init__(self, mgr, all_chars, config_path: Optional[str] = None, parent=None):
        super().__init__(parent)
        self.mgr = mgr
        self.all_chars = filter_out_benched(all_chars) if all_chars else []
        self.config_path = config_path
        self.setWindowTitle("📊 全账号百战技能总览")
        self.resize(1400, 800)
        icon = get_app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)

        from bz_core_skills import (
            load_core_skill_categories,
            get_best_candidate_skill,
            get_top_candidate_skills,
            get_core_skills_config_path,
            get_level_colors,
        )
        self._load_core_skill_categories = load_core_skill_categories
        self._get_best_candidate_skill = get_best_candidate_skill
        self._get_top_candidate_skills = get_top_candidate_skills
        self._get_core_skills_config_path = get_core_skills_config_path
        self.categories = self._load_core_skill_categories(self.config_path)
        self._level_colors = get_level_colors(self.config_path)

        self.setStyleSheet("""
            QDialog { background-color: #1e1e2d; color: #ffffff; font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; }
            QLabel { color: #eee; }
            QPushButton#PrimaryBtn { background-color: #0d47a1; color: white; font-weight: bold; border-radius: 4px; padding: 6px 14px; }
            QPushButton#PrimaryBtn:hover { background-color: #1565c0; }
            QPushButton#GreenBtn { background-color: #2e7d32; color: white; font-weight: bold; border-radius: 4px; padding: 6px 14px; }
            QPushButton#GreenBtn:hover { background-color: #388e3c; }
            QTableWidget { background-color: #161622; color: #ffffff; gridline-color: #2d2d3f; border: 1px solid #333345; border-radius: 6px; }
            QTableWidget::item { padding: 4px 6px; }
            QTableWidget::item:selected { background-color: #2b3b5c; }
            QHeaderView::section { background-color: #232336; color: #d0d4dc; font-weight: bold; border: 1px solid #333345; padding: 6px 4px; }
            QScrollBar:vertical { background: #161622; width: 10px; }
            QScrollBar::handle:vertical { background: #3b3b54; border-radius: 5px; }
            QScrollBar:horizontal { background: #161622; height: 10px; }
            QScrollBar::handle:horizontal { background: #3b3b54; border-radius: 5px; }
        """)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # 顶部工具栏
        top_bar = QHBoxLayout()
        top_bar.setSpacing(12)

        sorted_chars = sorted(self.all_chars, key=lambda c: c.get("name", ""))
        total_count = len(sorted_chars)
        has_bz_count = sum(
            1 for c in sorted_chars
            if c.get("baizhan_api") and isinstance(c.get("baizhan_api"), dict) and "error" not in c.get("baizhan_api") and c.get("baizhan_api").get("skillList")
        )

        self.lbl_stats = QLabel(f"共 <b>{total_count}</b> 个角色 | 有百战数据 <b>{has_bz_count}</b> 个")
        self.lbl_stats.setStyleSheet("font-size: 14px; color: #3b8ed0;")
        top_bar.addWidget(self.lbl_stats)

        top_bar.addStretch()

        self.btn_export_csv = QPushButton("⬇ 导出 CSV")
        self.btn_export_csv.setObjectName("GreenBtn")
        self.btn_export_csv.setToolTip("导出当前全账号百战总览表格为 CSV 文件")
        self.btn_export_csv.clicked.connect(self.export_to_csv)
        top_bar.addWidget(self.btn_export_csv)

        self.btn_config = QPushButton("⚙ 分类配置")
        self.btn_config.setObjectName("PrimaryBtn")
        self.btn_config.setToolTip("打开分类配置界面：选择分类、自定义分类、设置每类展示技能数量")
        self.btn_config.clicked.connect(self.open_config_dialog)
        top_bar.addWidget(self.btn_config)

        layout.addLayout(top_bar)

        # 表格初始化
        self.table = QTableWidget()
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)

        layout.addWidget(self.table)

        # 底部提示小字
        self.lbl_tip = QLabel("💡 技能分类支持在 ⚙ 分类配置 中自由定制（启用/禁用、新增分类、调整展示技能数量与优先级）。")
        self.lbl_tip.setStyleSheet("font-size: 12px; color: #8888aa; margin-top: 4px;")
        layout.addWidget(self.lbl_tip)

        # 构建表格内容
        self.rebuild_table()

    def open_config_dialog(self):
        dlg = CoreSkillsConfigDialog(config_path=self.config_path, parent=self)
        if dlg.exec() and getattr(dlg, "saved", False):
            self.categories = self._load_core_skill_categories(self.config_path)
            from bz_core_skills import get_level_colors
            self._level_colors = get_level_colors(self.config_path)
            self.rebuild_table()

    def rebuild_table(self):
        sorted_chars = sorted(self.all_chars, key=lambda c: c.get("name", ""))
        enabled_cats = [c for c in self.categories if c.get("enabled", True)]

        headers = ["角色", "服务器", "门派", "百战精", "百战耐"] + [
            f"{c.get('group', '')}·{c.get('window', '')}" for c in enabled_cats
        ]

        self.table.setSortingEnabled(False)
        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        # 列宽设置
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(1, 90)
        self.table.setColumnWidth(2, 80)
        self.table.setColumnWidth(3, 90)
        self.table.setColumnWidth(4, 90)
        for col_idx in range(5, len(headers)):
            self.table.setColumnWidth(col_idx, 130)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)

        if not enabled_cats:
            self.lbl_tip.setText("⚠️ 当前未启用任何技能分类，请点击 ⚙ 分类配置 启用")
            self.lbl_tip.setStyleSheet("font-size: 12px; color: #ff9800; margin-top: 4px;")
            self.table.verticalHeader().setDefaultSectionSize(26)
        else:
            max_disp = max([c.get("display_count", 1) for c in enabled_cats], default=1)
            row_height = 24 + max(0, max_disp - 1) * 16
            self.table.verticalHeader().setDefaultSectionSize(row_height)
            self.lbl_tip.setText("💡 技能分类支持在 ⚙ 分类配置 中自由定制（启用/禁用、新增分类、调整展示技能数量与优先级）。")
            self.lbl_tip.setStyleSheet("font-size: 12px; color: #8888aa; margin-top: 4px;")

        self.populate_table(sorted_chars, enabled_cats)
        self.table.setSortingEnabled(True)

    def populate_table(self, sorted_chars, enabled_cats):
        self.table.setRowCount(len(sorted_chars))

        for row_idx, c in enumerate(sorted_chars):
            name = c.get("name", "")
            server = c.get("server", "")
            sect = c.get("force", c.get("sect", c.get("school", "")))
            bz_api = c.get("baizhan_api", {})

            has_bz = bool(bz_api and isinstance(bz_api, dict) and "error" not in bz_api and bz_api.get("skillList"))

            # 1. 角色名
            it_name = QTableWidgetItem(name)
            it_name.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 0, it_name)

            # 2. 服务器
            it_srv = QTableWidgetItem(server)
            it_srv.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 1, it_srv)

            # 3. 门派
            it_sect = QTableWidgetItem(sect)
            it_sect.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 2, it_sect)

            # 4. 百战精
            stamina_val = bz_api.get("skillStamina") if (has_bz and isinstance(bz_api, dict)) else None
            if stamina_val is not None and isinstance(stamina_val, (int, float)):
                it_stamina = NumericTableWidgetItem(f"{stamina_val:,}", stamina_val)
                it_stamina.setForeground(QColor("#ffffff"))
            else:
                it_stamina = NumericTableWidgetItem("-", -1)
                it_stamina.setForeground(QColor("#777777"))
            it_stamina.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 3, it_stamina)

            # 5. 百战耐
            energy_val = bz_api.get("skillEnergy") if (has_bz and isinstance(bz_api, dict)) else None
            if energy_val is not None and isinstance(energy_val, (int, float)):
                it_energy = NumericTableWidgetItem(f"{energy_val:,}", energy_val)
                it_energy.setForeground(QColor("#ffffff"))
            else:
                it_energy = NumericTableWidgetItem("-", -1)
                it_energy.setForeground(QColor("#777777"))
            it_energy.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 4, it_energy)

            # 6~N. 技能列
            skill_list = bz_api.get("skillList", []) if has_bz else []

            for col_offset, cat in enumerate(enabled_cats):
                col_idx = 5 + col_offset
                candidates = cat.get("candidates", [])
                disp_count = max(1, int(cat.get("display_count", 1)))

                if not has_bz:
                    it_skill = NumericTableWidgetItem("无数据", -1)
                    it_skill.setForeground(QColor("#777777"))
                    it_skill.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    it_skill.setToolTip("该角色暂无本地百战招式数据")
                    self.table.setItem(row_idx, col_idx, it_skill)
                    continue

                if not candidates:
                    it_skill = NumericTableWidgetItem("—", -1)
                    it_skill.setForeground(QColor("#777777"))
                    it_skill.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    it_skill.setToolTip("该档位暂无归类技能，可点击 ⚙ 分类配置 自定义")
                    self.table.setItem(row_idx, col_idx, it_skill)
                    continue

                top_skills = self._get_top_candidate_skills(skill_list, candidates, top_n=disp_count)
                best_skill, learned_list = self._get_best_candidate_skill(skill_list, candidates)

                if not top_skills:
                    it_skill = NumericTableWidgetItem("未学习", 0)
                    it_skill.setForeground(QColor("#888888"))
                    it_skill.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    cand_text = ", ".join(candidates)
                    it_skill.setToolTip(f"未学习该档位技能\n\n【该档候选技能 ({len(candidates)}个)】\n{cand_text}")
                    self.table.setItem(row_idx, col_idx, it_skill)
                else:
                    lines = [f"{s.get('szSkillName', '')} Lv{int(s.get('nLevel', 0))}" for s in top_skills]
                    text = "\n".join(lines)
                    max_lvl = int(top_skills[0].get("nLevel", 0))
                    it_skill = NumericTableWidgetItem(text, max_lvl)
                    it_skill.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    # 等级颜色：按角色已学最高等级匹配规则（自上而下，命中即用）
                    color = None
                    for rule in self._level_colors:
                        try:
                            if int(rule.get("min", 1)) <= max_lvl <= int(rule.get("max", 999)):
                                color = rule.get("color")
                                break
                        except (ValueError, TypeError):
                            continue
                    if color:
                        it_skill.setForeground(QColor(color))
                    if max_lvl >= 10:
                        font = it_skill.font()
                        font.setBold(True)
                        it_skill.setFont(font)

                    # 构建 Tooltip
                    learned_lines = [f"{s.get('szSkillName', '')} Lv{s.get('nLevel', 0)}" for s in learned_list]
                    tooltip_lines = [
                        f"【已学候选 ({len(learned_list)}个)】",
                        "\n".join(learned_lines),
                        f"\n【该档全部候选 ({len(candidates)}个)】",
                        ", ".join(candidates)
                    ]
                    it_skill.setToolTip("\n".join(tooltip_lines))
                    self.table.setItem(row_idx, col_idx, it_skill)

    def export_to_csv(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, "导出百战总览 CSV", "baizhan_overview.csv", "CSV Files (*.csv)"
        )
        if not filename:
            return

        try:
            with open(filename, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                writer.writerow(headers)

                for row in range(self.table.rowCount()):
                    row_data = []
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)

            QMessageBox.information(self, "导出成功", f"全账号百战总览已成功导出至:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 CSV 遇到错误:\n{str(e)}")


class BaizhanSkillsDialog(QDialog):
    """百战招式面板独立弹窗"""
    def __init__(self, mgr, all_chars, default_char_name=None, parent=None):
        super().__init__(parent)
        self.mgr = mgr
        self.all_chars = filter_out_benched(all_chars) if all_chars else []
        self.current_bz_char_name = default_char_name
        self.skill_descs = self.load_skill_descriptions()
        self.active_color_filter = None
        self.active_type_filter = None

        self.setWindowTitle("⚔️ 百战异闻录 - 角色招式数据大局览")
        self.resize(1120, 760)
        self.setStyleSheet("""
            QDialog { background-color: #1e1e2d; color: #ffffff; font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; }
            QLabel { color: #eee; }
            QComboBox { background-color: #161622; color: #fff; border: 1px solid #444460; padding: 4px 8px; border-radius: 4px; }
            QCheckBox { color: #eee; font-weight: bold; }
            QPushButton#PrimaryBtn { background-color: #0d47a1; color: white; font-weight: bold; border-radius: 4px; padding: 6px 14px; }
            QPushButton#PrimaryBtn:hover { background-color: #1565c0; }
            QScrollArea { border: 1px solid #333345; background-color: #161622; border-radius: 6px; }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        # Header Bar
        header_bar = QHBoxLayout()
        header_bar.setSpacing(10)

        header_bar.addWidget(QLabel("<b>选择角色:</b>"))
        self.combo_char = QComboBox()
        self.combo_char.setMinimumWidth(160)
        char_names = [c.get("name", "") for c in self.all_chars if c.get("name")]
        self.combo_char.addItems(char_names)
        if default_char_name and default_char_name in char_names:
            self.combo_char.setCurrentText(default_char_name)
        elif char_names:
            self.current_bz_char_name = char_names[0]
        
        self.combo_char.currentTextChanged.connect(self.on_char_changed)
        header_bar.addWidget(self.combo_char)

        self.lbl_bz_char = QLabel("")
        self.lbl_bz_char.setStyleSheet("font-size: 13px; font-weight: bold; color: #3b8ed0;")
        header_bar.addWidget(self.lbl_bz_char)
        header_bar.addStretch()

        self.btn_all_overview = QPushButton("📊 全账号总览")
        self.btn_all_overview.setObjectName("PrimaryBtn")
        self.btn_all_overview.setToolTip("以表格展示全部角色的核心打精/打耐/回复技能等级")
        self.btn_all_overview.clicked.connect(self.open_all_accounts_overview)
        header_bar.addWidget(self.btn_all_overview)

        self.btn_refresh_bz_online = QPushButton("⟳ 强制在线刷新百战数据")
        self.btn_refresh_bz_online.setObjectName("PrimaryBtn")
        self.btn_refresh_bz_online.clicked.connect(self.refresh_current_char_bz_online)
        header_bar.addWidget(self.btn_refresh_bz_online)

        self.btn_export_excel = QPushButton("⬇ 导出 Excel")
        self.btn_export_excel.setObjectName("GreenBtn")
        self.btn_export_excel.setToolTip("将当前角色的所有百战招式数据导出为 Excel 表格 (.xlsx)")
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        header_bar.addWidget(self.btn_export_excel)

        self.btn_export_img = QPushButton("🖼️ 导出图片")
        self.btn_export_img.setObjectName("PurpleBtn")
        self.btn_export_img.setToolTip("导出当前筛选招式图或全技能长图")
        self.btn_export_img.clicked.connect(self.show_export_img_menu)
        header_bar.addWidget(self.btn_export_img)

        layout.addLayout(header_bar)

        # Filter Bar
        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(8)

        filter_bar.addWidget(QLabel("破绽颜色:"))
        self.combo_bz_color = QComboBox()
        self.combo_bz_color.addItems(["全部颜色", "黄破绽", "蓝破绽", "绿破绽", "红破绽", "紫破绽", "黑破绽", "白破绽"])
        self.combo_bz_color.currentTextChanged.connect(lambda _: self.re_render_bz())
        filter_bar.addWidget(self.combo_bz_color)

        filter_bar.addSpacing(6)
        filter_bar.addWidget(QLabel("招式类型:"))
        self.combo_bz_type = QComboBox()
        self.combo_bz_type.addItems(["全部类型", "攻击", "控制", "位移", "治疗", "特殊"])
        self.combo_bz_type.currentTextChanged.connect(lambda _: self.re_render_bz())
        filter_bar.addWidget(self.combo_bz_type)

        filter_bar.addSpacing(6)
        self.chk_bz_section = QCheckBox("仅看本周BOSS")
        self.chk_bz_section.setToolTip("勾选后限制仅显示本周排班及修罗 BOSS 的招式需求")
        self.chk_bz_section.stateChanged.connect(lambda _: self.re_render_bz())
        filter_bar.addWidget(self.chk_bz_section)

        self.combo_bz_section = QComboBox()
        self.combo_bz_section.addItems(["全部排班层数", "1 - 50 层首领", "50 - 70 层首领", "60 - 90 层首领", "70 - 90 层首领", "80 - 100 层首领", "90 - 100 层首领", "本周修罗/镇守首领"])
        self.combo_bz_section.setEnabled(False)
        self.chk_bz_section.toggled.connect(self.combo_bz_section.setEnabled)
        self.combo_bz_section.currentTextChanged.connect(lambda _: self.re_render_bz())
        filter_bar.addWidget(self.combo_bz_section)

        filter_bar.addSpacing(6)
        self.chk_bz_weekly_need = QCheckBox("仅看需求")
        self.chk_bz_weekly_need.setToolTip("勾选后仅展示当前角色尚未满级的招式需求")
        self.chk_bz_weekly_need.stateChanged.connect(lambda _: self.re_render_bz())
        filter_bar.addWidget(self.chk_bz_weekly_need)

        filter_bar.addSpacing(6)
        filter_bar.addWidget(QLabel("排序:"))
        self.combo_bz_sort = QComboBox()
        self.combo_bz_sort.addItems(["按首领名称", "按技能等级 (高→低)", "按技能等级 (低→高)", "按破绽颜色"])
        self.combo_bz_sort.currentTextChanged.connect(lambda _: self.re_render_bz())
        filter_bar.addWidget(self.combo_bz_sort)
        filter_bar.addStretch()

        layout.addLayout(filter_bar)

        # Scroll Area for Cards
        self.bz_scroll = QScrollArea()
        self.bz_scroll.setWidgetResizable(True)
        self.bz_container = QWidget()
        self.bz_layout = QVBoxLayout(self.bz_container)
        self.bz_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.bz_scroll.setWidget(self.bz_container)
        layout.addWidget(self.bz_scroll)

        # Initial Render
        target_name = self.current_bz_char_name or (char_names[0] if char_names else None)
        if target_name:
            self.render_bz_skills(target_name)

    def on_char_changed(self, name):
        if name:
            self.render_bz_skills(name)

    def refresh_current_char_bz_online(self):
        char_name = self.combo_char.currentText()
        if not char_name:
            return
        self.lbl_bz_char.setText(f"⌛ 正在向 JX3API 强制刷新 [{char_name}] 的最新百战数据...")
        self.btn_refresh_bz_online.setEnabled(False)
        self.btn_refresh_bz_online.setText("⟳ 正在在线刷新...")

        self.fetch_thread = ApiFetchThread(self.mgr, char_name)
        self.fetch_thread.fetched.connect(self.on_bz_fetched)
        self.fetch_thread.start()

    def on_bz_fetched(self, name, data):
        self.btn_refresh_bz_online.setEnabled(True)
        self.btn_refresh_bz_online.setText("⟳  强制在线刷新百战数据")

        if "error" in data:
            QMessageBox.warning(self, "错误", f"在线刷新失败: {data['error']}")
            return

        c = self.mgr.characters.get(name)
        if c:
            c["baizhan_api"] = data

        self.render_bz_skills(name)

    def re_render_bz(self):
        char_name = self.combo_char.currentText() if hasattr(self, "combo_char") else getattr(self, "current_bz_char_name", None)
        if char_name:
            self.render_bz_skills(char_name)

    def load_skill_descriptions(self):
        desc_path = os.path.join(os.path.dirname(__file__), "data", "bz_skill_desc.json")
        if os.path.exists(desc_path):
            try:
                with open(desc_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except: pass
        return {}

    def format_skill_desc_for_level(self, desc_data, level):
        if isinstance(desc_data, dict):
            text = desc_data.get("detail", "") or desc_data.get("brief", "")
        else:
            text = str(desc_data) if desc_data else ""

        if not text:
            return "暂无该招式详细说明。"

        target_idx = max(0, min(9, (level - 1) if level > 0 else 0))

        def repl(match):
            raw_seq = match.group(1)
            parts = raw_seq.split("/")
            if len(parts) >= 10:
                val = parts[target_idx].strip()
                return f"<b style='color: #ffe066;'>{val}</b>"
            elif len(parts) > 1:
                idx = min(target_idx, len(parts) - 1)
                val = parts[idx].strip()
                return f"<b style='color: #ffe066;'>{val}</b>"
            return match.group(0)

        import re
        formatted = re.sub(r"\[([0-9\.,/]+)\]", repl, text)
        return formatted

    def format_skill_desc_plain_text(self, desc_data, level):
        if isinstance(desc_data, dict):
            text = desc_data.get("detail", "") or desc_data.get("brief", "")
        else:
            text = str(desc_data) if desc_data else ""

        if not text:
            return "暂无说明"

        target_idx = max(0, min(9, (level - 1) if level > 0 else 0))

        def repl(match):
            raw_seq = match.group(1)
            parts = raw_seq.split("/")
            if len(parts) >= 10:
                return parts[target_idx].strip()
            elif len(parts) > 1:
                idx = min(target_idx, len(parts) - 1)
                return parts[idx].strip()
            return match.group(0)

        import re
        plain = re.sub(r"\[([0-9\.,/]+)\]", repl, text)
        return plain

    def export_to_excel(self):
        char_name = self.combo_char.currentText().strip()
        if not char_name:
            QMessageBox.warning(self, "导出提示", "未选中有效的角色！")
            return

        c = next((x for x in self.all_chars if x.get("name") == char_name), None)
        if not c:
            c = self.mgr.characters.get(char_name, {})

        bz_api = c.get("baizhan_api", {}) or {}
        skill_list = bz_api.get("skillList", []) or []

        if not skill_list:
            QMessageBox.warning(self, "导出提示", f"角色【{char_name}】暂无百战招式数据！")
            return

        filePath, _ = QFileDialog.getSaveFileName(
            self,
            "导出 Excel 表格",
            f"【{char_name}】_百战招式统计.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filePath:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "百战招式统计"

            # Title Row
            ws.merge_cells("A1:H1")
            title_cell = ws["A1"]
            title_cell.value = f"【{char_name}】百战异闻录 招式统计汇总"
            title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40

            # Meta Row
            stamina = bz_api.get("skillStamina", "-")
            energy = bz_api.get("skillEnergy", "-")
            update_time = bz_api.get("updateTime", "未知")
            ws.merge_cells("A2:H2")
            meta_cell = ws["A2"]
            meta_cell.value = f"角色精耐: 精力 {stamina} / 耐力 {energy}    |    数据同步时间: {update_time}    |    统计招式总数: {len(skill_list)} 个"
            meta_cell.font = Font(name="微软雅黑", size=11, italic=True, color="D9D9D9")
            meta_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            meta_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 24

            # Table Headers
            headers = ["招式名称", "重数", "破绽颜色", "招式类型", "来源首领", "本周需求/排班层数", "需求状态", "招式详细说明"]
            ws.append([]) # Blank row 3
            ws.append(headers) # Row 4
            ws.row_dimensions[4].height = 28

            header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            header_border = Border(
                left=Side(style="thin", color="808080"),
                right=Side(style="thin", color="808080"),
                top=Side(style="thin", color="808080"),
                bottom=Side(style="medium", color="000000")
            )

            for col_num in range(1, 9):
                cell = ws.cell(row=4, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = header_border

            # Roster Map for Demand floor calculation
            wb_data = getattr(self.mgr, "weekly_bosses", {}) or {}
            boss_list = wb_data.get("list", []) if isinstance(wb_data, dict) and "list" in wb_data else []
            custom_xiuluo = getattr(self.mgr, "custom_xiuluo_boss", None)
            xiuluo_raw = custom_xiuluo if custom_xiuluo else (wb_data.get("boss", "") if isinstance(wb_data, dict) else "")
            xiuluo_bosses = {t.strip() for t in xiuluo_raw.replace("／", "/").split("/") if t.strip()}

            def get_floor_drop_level(idx):
                if idx <= 0: return 10
                return 10 if idx >= 91 else min(10, (idx - 1) // 10 + 1)

            boss_floors_map = {}
            for item in boss_list:
                raw_bname = item.get("name", "").strip()
                idx = item.get("index", 0)
                if idx > 0 and raw_bname:
                    aliases = get_boss_aliases(raw_bname)
                    for alias in aliases:
                        if alias not in boss_floors_map: boss_floors_map[alias] = []
                        if idx not in boss_floors_map[alias]: boss_floors_map[alias].append(idx)

            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9")
            )

            # Data Rows
            sorted_sks = sorted(skill_list, key=lambda s: (-s.get("nLevel", 0), s.get("nColor", 0)))
            for row_idx, sk in enumerate(sorted_sks, start=5):
                sname = sk.get("szSkillName") or sk.get("szName") or "未知"
                slvl = sk.get("nLevel", 0)
                scol = sk.get("nColor", 0)
                sboss = (sk.get("szBossName") or sk.get("boss") or "通用/未知").strip()
                if sboss.startswith("恶战") or "恶战" in sboss:
                    sboss = "恶战"
                zt = sk.get("szType", "")

                ts = set()
                if zt:
                    for t in zt.split(";"):
                        ts.add(TYPE_MAP.get(t, t))
                tstr = ",".join(sorted(ts)) if ts else "-"

                cinfo = COLOR_ORDER.get(scol, (99, "其它"))
                color_name = cinfo[1]

                is_max = "已满重" if slvl >= 10 else f"需提升({slvl}➜10重)"

                floors_for_b = get_floors_for_skill_boss(sboss, boss_floors_map)
                useful_f = [f for f in floors_for_b if get_floor_drop_level(f) > slvl]
                if sboss in xiuluo_bosses:
                    floors_str = "修罗首领"
                elif useful_f:
                    floors_str = "、".join(f"{f}层" for f in useful_f)
                else:
                    floors_str = "本周未排班/无满足"

                s_desc_data = self.skill_descs.get(sname, {})
                plain_desc = self.format_skill_desc_plain_text(s_desc_data, slvl)

                row_data = [sname, slvl, color_name, tstr, sboss, floors_str, is_max, plain_desc]
                ws.append(row_data)

                ws.row_dimensions[row_idx].height = 22

                row_fill = PatternFill(start_color="F9F9F9" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")

                for c_i in range(1, 9):
                    c_cell = ws.cell(row=row_idx, column=c_i)
                    c_cell.font = Font(name="微软雅黑", size=10)
                    c_cell.fill = row_fill
                    c_cell.border = thin_border
                    align_h = "center" if c_i in (2, 3, 4, 5, 6, 7) else "left"
                    c_cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)

            # Auto Column Widths
            col_widths = {1: 18, 2: 10, 3: 12, 4: 12, 5: 16, 6: 22, 7: 16, 8: 45}
            for col_i, width in col_widths.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = width

            wb.save(filePath)
            QMessageBox.information(self, "导出成功", f"✓ 已成功导出角色【{char_name}】的百战招式数据至:\n\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 Excel 表格时发生异常:\n{e}")

    def show_export_img_menu(self):
        char_name = self.combo_char.currentText().strip()
        if not char_name:
            QMessageBox.warning(self, "导出提示", "未选中有效的角色！")
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #1e1e32;
                color: #ffffff;
                border: 1px solid #4a4a7a;
                border-radius: 6px;
                padding: 4px;
            }
            QMenu::item {
                padding: 8px 20px;
                border-radius: 4px;
            }
            QMenu::item:selected {
                background-color: #3b63ce;
                color: #ffffff;
                font-weight: bold;
            }
        """)

        act_filtered = menu.addAction("🎯 导出当前筛选视图图片")
        act_all = menu.addAction("📜 导出所有技能长图 (忽略筛选)")

        action = menu.exec(QCursor.pos())
        if action == act_filtered:
            self.export_skills_image(export_mode="filtered")
        elif action == act_all:
            self.export_skills_image(export_mode="all")

    def export_skills_image(self, export_mode="filtered"):
        char_name = self.combo_char.currentText().strip()
        if not char_name:
            return

        c = next((x for x in self.all_chars if x.get("name") == char_name), None)
        if not c:
            c = self.mgr.characters.get(char_name, {})

        skill_list = c.get("baizhan_api", {}).get("skillList", []) if c else []
        if not skill_list:
            QMessageBox.warning(self, "导出提示", "暂无招式数据可导出！")
            return

        if export_mode == "filtered":
            skills_to_export = getattr(self, "current_filtered_skills", skill_list)
            title_tag = "当前筛选"
        else:
            skills_to_export = skill_list
            title_tag = "全技能汇总"

        if not skills_to_export:
            QMessageBox.warning(self, "导出提示", "符合条件的导出招式为空！")
            return

        default_name = f"【{char_name}】_百战招式_{title_tag}.png"
        filePath, _ = QFileDialog.getSaveFileName(
            self,
            "导出招式图片",
            default_name,
            "PNG Images (*.png);;JPEG Images (*.jpg)"
        )

        if filePath:
            export_w = CompactSkillsExportWidget(char_name, skills_to_export, title_tag=title_tag)
            export_w.adjustSize()
            QApplication.processEvents()

            pixmap = export_w.grab()
            success = pixmap.save(filePath)
            export_w.deleteLater()

            if success:
                QMessageBox.information(self, "导出成功", f"✓ 已成功导出【{char_name}】的【{title_tag}】精简长图至:\n\n{filePath}")
            else:
                QMessageBox.critical(self, "导出失败", "保存图片文件失败，请检查写入权限！")

    def render_bz_skills(self, name):
        self.current_bz_char_name = name
        c = next((x for x in self.all_chars if x.get("name") == name), None)
        if not c:
            c = self.mgr.characters.get(name, {})

        bz_api = c.get("baizhan_api", {}) if c else {}
        if not bz_api and c and c.get("server"):
            from readers.baizhan_api import api as bz_api_module
            bz_api = bz_api_module.get_character_skills(c.get("server"), name, force_refresh=False)
            if bz_api and "error" not in bz_api:
                c["baizhan_api"] = bz_api

        fetch_time = bz_api.get("_fetch_time", "未同步") if isinstance(bz_api, dict) else "未同步"
        api_date = bz_api.get("_api_date", bz_api.get("date", fetch_time.split()[0] if " " in fetch_time else "未知")) if isinstance(bz_api, dict) else "未知"
        
        self.lbl_bz_char.setText(f"⚔️ {name} 的百战招式  [API数据日期: {api_date} | 客户端本地同步时间: {fetch_time}]")

        # Clear existing layout
        while self.bz_layout.count():
            child = self.bz_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if not bz_api or "error" in bz_api or not bz_api.get("skillList"):
            lbl = QLabel("暂无该角色的本地百战数据缓存。可点击右侧“⟳ 强制在线刷新百战数据”进行在线同步。")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("font-size: 14px; color: #aaa; margin-top: 40px;")
            self.bz_layout.addWidget(lbl)
            return

        skill_list = c.get("baizhan_api", {}).get("skillList", [])
        if not skill_list:
            lbl = QLabel("该角色暂未学习百战招式。")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.bz_layout.addWidget(lbl)
            return

        sort_mode = self.combo_bz_sort.currentText() if hasattr(self, "combo_bz_sort") else "按首领名称"

        COLOR_DROPDOWN_MAP = {
            "黄破绽": 2, "蓝破绽": 3, "绿破绽": 4, "红破绽": 5, "紫破绽": 6, "黑破绽": 7, "白破绽": 0
        }
        sel_col_text = self.combo_bz_color.currentText() if hasattr(self, "combo_bz_color") else "全部颜色"
        target_color = COLOR_DROPDOWN_MAP.get(sel_col_text, None)

        sel_type_text = self.combo_bz_type.currentText() if hasattr(self, "combo_bz_type") else "全部类型"
        target_type = None if sel_type_text == "全部类型" else sel_type_text

        use_section = self.chk_bz_section.isChecked() if hasattr(self, "chk_bz_section") else False
        sec_text = self.combo_bz_section.currentText() if (use_section and hasattr(self, "combo_bz_section")) else "全部排班层数"
        only_need = self.chk_bz_weekly_need.isChecked() if hasattr(self, "chk_bz_weekly_need") else False

        wb = getattr(self.mgr, "weekly_bosses", {}) or {}
        boss_list = wb.get("list", []) if isinstance(wb, dict) and "list" in wb else []
        custom_xiuluo = getattr(self.mgr, "custom_xiuluo_boss", None)
        xiuluo_raw = custom_xiuluo if custom_xiuluo else (wb.get("boss", "") if isinstance(wb, dict) else "")
        xiuluo_bosses = {t.strip() for t in xiuluo_raw.replace("／", "/").split("/") if t.strip()}

        def get_floor_drop_level(idx):
            if idx <= 0: return 10
            return 10 if idx >= 91 else min(10, (idx - 1) // 10 + 1)

        def is_in_selected_section(f):
            if not use_section or sec_text == "全部排班层数": return True
            if sec_text == "1 - 50 层首领": return 1 <= f <= 50
            if sec_text == "50 - 70 层首领": return 50 <= f <= 70
            if sec_text == "60 - 90 层首领": return 60 <= f <= 90
            if sec_text == "70 - 90 层首领": return 70 <= f <= 90
            if sec_text == "80 - 100 层首领": return 80 <= f <= 100
            if sec_text == "90 - 100 层首领": return 90 <= f <= 100
            if sec_text == "本周修罗/镇守首领": return False
            return True

        boss_target_drop_map = {}
        boss_floors_map = {}

        for item in boss_list:
            raw_bname = item.get("name", "").strip()
            idx = item.get("index", 0)
            dlvl = get_floor_drop_level(idx)

            if idx > 0 and raw_bname:
                aliases = get_boss_aliases(raw_bname)
                for alias in aliases:
                    if alias not in boss_floors_map: boss_floors_map[alias] = []
                    if idx not in boss_floors_map[alias]: boss_floors_map[alias].append(idx)

            in_sec = False
            if sec_text == "1 - 50 层首领" and 1 <= idx <= 50: in_sec = True
            elif sec_text == "50 - 70 层首领" and 50 <= idx <= 70: in_sec = True
            elif sec_text == "60 - 90 层首领" and 60 <= idx <= 90: in_sec = True
            elif sec_text == "70 - 90 层首领" and 70 <= idx <= 90: in_sec = True
            elif sec_text == "80 - 100 层首领" and 80 <= idx <= 100: in_sec = True
            elif sec_text == "90 - 100 层首领" and 90 <= idx <= 100: in_sec = True
            elif sec_text == "本周修罗/镇守首领" and any(a in xiuluo_bosses for a in get_boss_aliases(raw_bname)): in_sec = True
            elif sec_text == "全部排班层数": in_sec = True

            if in_sec and raw_bname:
                if sec_text == "本周修罗/镇守首领":
                    dlvl = 10
                aliases = get_boss_aliases(raw_bname)
                for alias in aliases:
                    boss_target_drop_map[alias] = max(boss_target_drop_map.get(alias, 0), dlvl)
                    if alias in ANOMALY_MAP:
                        base_b = ANOMALY_MAP[alias]
                        boss_target_drop_map[base_b] = max(boss_target_drop_map.get(base_b, 0), dlvl)

        for xb in xiuluo_bosses:
            if sec_text in ("全部排班层数", "本周修罗/镇守首领"):
                boss_target_drop_map[xb] = 10
                if xb in ANOMALY_MAP:
                    base_b = ANOMALY_MAP[xb]
                    boss_target_drop_map[base_b] = 10

        boss_skills_map = {}
        for sk in skill_list:
            raw_b = sk.get("szBossName") or sk.get("boss") or "通用/未知首领"
            boss = raw_b.strip()
            if boss.startswith("恶战") or "恶战" in boss:
                boss = "恶战"
            if boss not in boss_skills_map:
                boss_skills_map[boss] = []
            boss_skills_map[boss].append(sk)

        needed_bosses = set()
        if only_need:
            for boss_name, sks in boss_skills_map.items():
                req_lvl = boss_target_drop_map.get(boss_name)
                if req_lvl is None:
                    if use_section:
                        req_lvl = 0
                    else:
                        req_lvl = 10 if boss_name in xiuluo_bosses else 7

                if req_lvl > 0 and any(sk.get("nLevel", 0) < req_lvl for sk in sks):
                    needed_bosses.add(boss_name)

        valid_items = []
        for sk in skill_list:
            raw_b = sk.get("szBossName") or sk.get("boss") or "通用/未知首领"
            boss = raw_b.strip()
            if boss.startswith("恶战") or "恶战" in boss:
                boss = "恶战"
            col = sk.get("nColor", 0)
            zt = sk.get("szType", "")
            lv = sk.get("nLevel", 0)

            if target_color is not None and col != target_color:
                continue

            ts = set()
            if zt:
                for t in zt.split(";"):
                    ts.add(TYPE_MAP.get(t, t))
            tstr = ",".join(sorted(ts)) if ts else "-"

            if target_type and target_type not in tstr:
                continue

            if use_section:
                if boss not in boss_target_drop_map:
                    continue

            if only_need:
                req_lvl = boss_target_drop_map.get(boss)
                if req_lvl is None:
                    if use_section:
                        req_lvl = 0
                    else:
                        req_lvl = 10 if boss in xiuluo_bosses else 7

                if req_lvl <= 0 or lv >= req_lvl:
                    continue

            valid_items.append((sk, tstr, boss, lv, col))

        self.current_filtered_skills = [x[0] for x in valid_items]

        from collections import OrderedDict
        grouped = OrderedDict()

        if sort_mode == "按技能等级 (高→低)":
            valid_items.sort(key=lambda x: (-x[3], COLOR_ORDER.get(x[4], (99, ""))[0], x[2]))
            for sk, tstr, boss, lv, col in valid_items:
                gkey = f"⭐ {lv} 重招式"
                if gkey not in grouped:
                    grouped[gkey] = {"info": None, "skills": []}
                grouped[gkey]["skills"].append((sk, tstr))

        elif sort_mode == "按技能等级 (低→高)":
            valid_items.sort(key=lambda x: (x[3], COLOR_ORDER.get(x[4], (99, ""))[0], x[2]))
            for sk, tstr, boss, lv, col in valid_items:
                gkey = f"⭐ {lv} 重招式"
                if gkey not in grouped:
                    grouped[gkey] = {"info": None, "skills": []}
                grouped[gkey]["skills"].append((sk, tstr))

        elif sort_mode == "按破绽颜色":
            valid_items.sort(key=lambda x: (COLOR_ORDER.get(x[4], (99, ""))[0], -x[3], x[2]))
            for sk, tstr, boss, lv, col in valid_items:
                cinfo = COLOR_ORDER.get(col, (99, "其它破绽"))
                gkey = f"🎨 {cinfo[1]}招式"
                if gkey not in grouped:
                    grouped[gkey] = {"info": None, "skills": []}
                grouped[gkey]["skills"].append((sk, tstr))

        else:  # 按首领名称
            def get_boss_sort_key(x):
                sk, tstr, boss, lv, col = x
                aliases = get_boss_aliases(boss)
                is_xiuluo = any(a in xiuluo_bosses for a in aliases)
                floors = [f for f in get_floors_for_skill_boss(boss, boss_floors_map) if is_in_selected_section(f)]
                
                sks_for_boss = []
                for a in aliases:
                    sks_for_boss.extend(boss_skills_map.get(a, []))
                boss_levels = [s.get("nLevel", 0) for s in sks_for_boss] if sks_for_boss else [lv]
                
                if only_need:
                    dfloors = [f for f in floors if any(l < get_floor_drop_level(f) for l in boss_levels)]
                else:
                    dfloors = floors
                
                if is_xiuluo:
                    first_floor = 9999
                elif dfloors:
                    first_floor = min(dfloors)
                elif floors:
                    first_floor = min(floors)
                else:
                    first_floor = 999
                
                return (first_floor, is_xiuluo, boss, -lv, COLOR_ORDER.get(col, (99, ""))[0])

            valid_items.sort(key=get_boss_sort_key)

            for sk, tstr, boss, lv, col in valid_items:
                if boss not in grouped:
                    aliases = get_boss_aliases(boss)
                    floors = get_floors_for_skill_boss(boss, boss_floors_map)
                    sks_for_boss = []
                    for a in aliases:
                        sks_for_boss.extend(boss_skills_map.get(a, []))
                    boss_levels = [s.get("nLevel", 0) for s in sks_for_boss] if sks_for_boss else [lv]

                    if only_need:
                        display_floors = [f for f in floors if is_in_selected_section(f) and any(l < get_floor_drop_level(f) for l in boss_levels)]
                    else:
                        display_floors = [f for f in floors if is_in_selected_section(f)]

                    b_info = {
                        "name": boss,
                        "floors": display_floors,
                        "is_xiuluo": any(a in xiuluo_bosses for a in aliases),
                        "target_drop": max([boss_target_drop_map.get(a, 0) for a in aliases] + [10 if any(a in xiuluo_bosses for a in aliases) else 7])
                    }
                    grouped[boss] = {"info": b_info, "skills": []}
                grouped[boss]["skills"].append((sk, tstr))

        if not grouped:
            lbl = QLabel("未找到匹配筛选条件的招式。")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("font-size: 14px; color: #888; margin-top: 30px;")
            self.bz_layout.addWidget(lbl)
            return

        for group_key, gdata in grouped.items():
            b_info = gdata["info"]
            sks = gdata["skills"]

            gbox = QGroupBox()
            gb_layout = QVBoxLayout(gbox)
            gb_layout.setContentsMargins(10, 10, 10, 10)
            gb_layout.setSpacing(6)

            if b_info:
                bname = b_info["name"]
                floors = b_info["floors"]
                is_x = b_info["is_xiuluo"]
                tdrop = b_info["target_drop"]

                header_widget = QWidget()
                h_lay = QHBoxLayout(header_widget)
                h_lay.setContentsMargins(0, 0, 0, 0)
                h_lay.setSpacing(6)

                lbl_bname = QLabel(f"<b>【{bname}】</b>")
                lbl_bname.setStyleSheet("font-size: 15px; font-weight: bold; color: #ffca28;")
                h_lay.addWidget(lbl_bname)

                if is_x:
                    lbl_work = QLabel("⚡ 本周上班 (修罗首领)")
                    lbl_work.setStyleSheet("background-color: #8e24aa; color: #ffffff; font-weight: bold; padding: 3px 8px; border-radius: 4px; font-size: 12px;")
                    h_lay.addWidget(lbl_work)
                elif floors:
                    lbl_work = QLabel(f"✅ 本周上班 ({len(floors)}个层数)")
                    lbl_work.setStyleSheet("background-color: #2e7d32; color: #ffffff; font-weight: bold; padding: 3px 8px; border-radius: 4px; font-size: 12px;")
                    h_lay.addWidget(lbl_work)
                else:
                    lbl_work = QLabel("☕ 本周不上班")
                    lbl_work.setStyleSheet("background-color: #424242; color: #aaaaaa; padding: 3px 8px; border-radius: 4px; font-size: 12px;")
                    h_lay.addWidget(lbl_work)

                h_lay.addStretch()

                if floors:
                    fl_str = "、".join(f"{f}层" for f in floors)
                    tag_title = "📍 需求排班" if only_need else "📍 本周上班"
                    lbl_floors = QLabel(f"{tag_title}: <b>{fl_str}</b>  (最高需 {tdrop} 重)")
                    lbl_floors.setStyleSheet("color: #ffb74d; font-size: 13px;")
                    h_lay.addWidget(lbl_floors)
                elif is_x:
                    tag_title = "📍 需求排班" if only_need else "📍 本周上班"
                    lbl_floors = QLabel(f"{tag_title}: <b>修罗首领</b>  (最高需 10 重)")
                    lbl_floors.setStyleSheet("color: #e1bee7; font-size: 13px;")
                    h_lay.addWidget(lbl_floors)

                gb_layout.addWidget(header_widget)
            else:
                lbl_gkey = QLabel(f"<span style='font-size: 15px; color: #64b5f6;'><b>{group_key}</b></span>")
                gb_layout.addWidget(lbl_gkey)

            grid = QGridLayout()
            grid.setSpacing(8)

            COLS = 3
            for idx, (sk, tstr) in enumerate(sks):
                r = idx // COLS
                col_idx = idx % COLS

                sname = sk.get("szSkillName") or sk.get("szName") or "未知技能"
                slvl = sk.get("nLevel", 0)
                scol = sk.get("nColor", 0)
                sboss = (sk.get("szBossName") or sk.get("boss") or "通用/未知首领").strip()
                if sboss.startswith("恶战") or "恶战" in sboss:
                    sboss = "恶战"
                iid = sk.get("dwInSkillID", 0)

                cinfo = COLOR_ORDER.get(scol, (99, "其它"))
                bg_c, border_c, text_c, label_c = COLOR_STYLES.get(scol, ("#2a2a3c", "#444460", "#ffffff", "#ffffff"))

                card = SkillCardWidget(bg_c, border_c)
                c_main_layout = QHBoxLayout(card)
                c_main_layout.setContentsMargins(8, 6, 8, 6)
                c_main_layout.setSpacing(8)

                # Icon
                lbl_icon = QLabel()
                lbl_icon.setFixedSize(42, 42)
                
                ld = os.path.join(os.path.dirname(__file__), "web", "icons")
                cache_dir = os.path.join(os.path.dirname(__file__), "data", "bz_cache", "icons")
                icon_fp = os.path.join(ld, f"{sname}.png")
                if not os.path.exists(icon_fp) and iid:
                    icon_fp = os.path.join(cache_dir, f"{iid}.png")

                if os.path.exists(icon_fp):
                    pix = QPixmap(icon_fp)
                    if not pix.isNull():
                        lbl_icon.setPixmap(pix.scaled(42, 42, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                        lbl_icon.setStyleSheet(f"border: 1px solid {border_c}; border-radius: 4px;")
                    else:
                        lbl_icon.setText(cinfo[1][:1])
                        lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        lbl_icon.setStyleSheet(f"background-color: {label_c}; color: #000; font-weight: bold; border-radius: 4px; font-size: 14px;")
                else:
                    lbl_icon.setText(cinfo[1][:1])
                    lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    lbl_icon.setStyleSheet(f"background-color: {label_c}; color: #000; font-weight: bold; border-radius: 4px; font-size: 14px;")

                c_main_layout.addWidget(lbl_icon)

                # Text Content Layout
                c_layout = QVBoxLayout()
                c_layout.setContentsMargins(0, 0, 0, 0)
                c_layout.setSpacing(2)

                top_l = QHBoxLayout()
                lbl_sn = QLabel(f"<b>{sname}</b>")
                lbl_sn.setStyleSheet(f"color: {text_c}; font-size: 13px;")
                top_l.addWidget(lbl_sn)
                top_l.addStretch()

                lbl_lv = QLabel(f"<b>{slvl} 重</b>")
                lbl_lv.setStyleSheet("color: #ffd54f; font-size: 13px; font-weight: bold;")
                top_l.addWidget(lbl_lv)
                c_layout.addLayout(top_l)

                bot_l = QHBoxLayout()
                lbl_tp = QLabel(f"类型: {tstr}")
                lbl_tp.setStyleSheet("color: #aaa; font-size: 11px;")
                bot_l.addWidget(lbl_tp)

                if sort_mode != "按首领名称":
                    lbl_b = QLabel(f"首领: {sboss}")
                    lbl_b.setStyleSheet("color: #aaa; font-size: 11px; margin-left: 6px;")
                    bot_l.addWidget(lbl_b)

                bot_l.addStretch()

                floors_for_b = get_floors_for_skill_boss(sboss, boss_floors_map)
                if use_section:
                    valid_sec_f = [f for f in floors_for_b if is_in_selected_section(f)]
                else:
                    valid_sec_f = floors_for_b

                if only_need:
                    if slvl < 10:
                        useful_f = [f for f in valid_sec_f if get_floor_drop_level(f) > slvl]
                        if sboss in xiuluo_bosses:
                            demand_str = "修罗首领"
                            demand_color = "#e1bee7"
                        elif useful_f:
                            demand_str = "、".join(f"{f}层" for f in useful_f)
                            demand_color = "#ffb74d"
                        else:
                            demand_str = "本周无满足"
                            demand_color = "#888888"

                        lbl_demand_info = QLabel(f"📍 需求: <font color='{demand_color}'><b>{demand_str}</b></font>")
                        lbl_demand_info.setStyleSheet("font-size: 11px; color: #b0bec5;")
                        bot_l.addWidget(lbl_demand_info)
                else:
                    if sboss in xiuluo_bosses:
                        work_str = "修罗首领"
                        work_color = "#e1bee7"
                    elif valid_sec_f:
                        work_str = "、".join(f"{f}层" for f in valid_sec_f)
                        work_color = "#81c784"
                    else:
                        work_str = "未排班"
                        work_color = "#888888"

                    lbl_work_info = QLabel(f"📍 上班: <font color='{work_color}'><b>{work_str}</b></font>")
                    lbl_work_info.setStyleSheet("font-size: 11px; color: #b0bec5;")
                    bot_l.addWidget(lbl_work_info)

                c_layout.addLayout(bot_l)

                c_main_layout.addLayout(c_layout)

                s_desc_data = self.skill_descs.get(sname, {})
                s_desc_formatted = self.format_skill_desc_for_level(s_desc_data, slvl)
                card.setToolTip(f"<b>{sname}</b> ({slvl}重)<br/>破绽: {cinfo[1]} | 类型: {tstr}<br/>来源首领: {sboss}<hr/>{s_desc_formatted}")

                grid.addWidget(card, r, col_idx)

            gb_layout.addLayout(grid)
            self.bz_layout.addWidget(gbox)

    def open_all_accounts_overview(self):
        """打开全账号百战技能总览弹窗"""
        active_chars = filter_out_benched(self.all_chars)
        dlg = AllAccountsBaizhanDialog(self.mgr, active_chars, parent=self)
        dlg.exec()


class BenchManagerDialog(QDialog):
    """🪑 待选区管理弹窗"""
    def __init__(self, mgr, parent=None):
        super().__init__(parent)
        self.mgr = mgr
        self.setWindowTitle("🪑 待选区管理")
        self.resize(460, 520)
        self.setStyleSheet(DARK_QSS)
        icon = get_app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)
        self.init_ui()
        self.load_bench_list()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        lbl_tip = QLabel("待选区角色不参与所有表格统计、副本汇总及百战总览。\n可在此集中查看并将角色移出待选区。")
        lbl_tip.setStyleSheet("color: #b0bec5; font-size: 12px; line-height: 1.4;")
        layout.addWidget(lbl_tip)

        self.list_bench = QListWidget()
        self.list_bench.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.list_bench.setStyleSheet("""
            QListWidget {
                background-color: #1e1e24;
                border: 1px solid #3f3f46;
                border-radius: 6px;
                padding: 6px;
                font-size: 13px;
                color: #e0e0e0;
            }
            QListWidget::item {
                padding: 6px 10px;
                border-radius: 4px;
                margin-bottom: 2px;
            }
            QListWidget::item:hover {
                background-color: #2a2a35;
            }
            QListWidget::item:selected {
                background-color: #3b8ed0;
                color: #ffffff;
            }
        """)
        layout.addWidget(self.list_bench)

        self.lbl_empty = QLabel("待选区为空。\n可在角色表格中右键角色将其移入待选区。")
        self.lbl_empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_empty.setStyleSheet("color: #777788; font-size: 13px; padding: 40px;")
        self.lbl_empty.setVisible(False)
        layout.addWidget(self.lbl_empty)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self.btn_remove_selected = QPushButton("↩ 移出选中")
        self.btn_remove_selected.setObjectName("PrimaryBtn")
        self.btn_remove_selected.clicked.connect(self.remove_selected_chars)
        btn_layout.addWidget(self.btn_remove_selected)

        self.btn_remove_all = QPushButton("↩ 全部移出")
        self.btn_remove_all.setObjectName("DangerBtn")
        self.btn_remove_all.setStyleSheet("""
            QPushButton {
                background-color: #7b1fa2; color: #ffffff; font-weight: bold;
                border: 1px solid #9c27b0; border-radius: 4px; padding: 5px 12px;
            }
            QPushButton:hover { background-color: #8e24aa; }
        """)
        self.btn_remove_all.clicked.connect(self.remove_all_chars)
        btn_layout.addWidget(self.btn_remove_all)

        btn_layout.addStretch()

        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)

        layout.addLayout(btn_layout)

    def load_bench_list(self):
        self.list_bench.clear()
        bench_mgr = getattr(self.mgr, "bench_mgr", None)
        benched_names = bench_mgr.get_all() if bench_mgr else []

        self.setWindowTitle(f"🪑 待选区管理 ({len(benched_names)}人)")
        if benched_names:
            self.list_bench.setVisible(True)
            self.lbl_empty.setVisible(False)
            self.btn_remove_selected.setEnabled(True)
            self.btn_remove_all.setEnabled(True)
            for name in benched_names:
                item = QListWidgetItem(f"🪑 {name}")
                item.setData(Qt.ItemDataRole.UserRole, name)
                self.list_bench.addItem(item)
        else:
            self.list_bench.setVisible(False)
            self.lbl_empty.setVisible(True)
            self.btn_remove_selected.setEnabled(False)
            self.btn_remove_all.setEnabled(False)

    def remove_selected_chars(self):
        selected_items = self.list_bench.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "提示", "请先在列表中选中要移出的角色。")
            return

        for item in selected_items:
            name = item.data(Qt.ItemDataRole.UserRole) or item.text().replace("🪑 ", "").strip()
            self.mgr.set_char_benched(name, False)

        if self.parent() and hasattr(self.parent(), "apply_filters"):
            self.parent().apply_filters()
            if hasattr(self.parent(), "statusBar"):
                self.parent().statusBar().showMessage(f"已将 {len(selected_items)} 个角色移出待选区", 4000)

        self.load_bench_list()

    def remove_all_chars(self):
        bench_mgr = getattr(self.mgr, "bench_mgr", None)
        benched_names = bench_mgr.get_all() if bench_mgr else []
        if not benched_names:
            return

        reply = QMessageBox.question(
            self, "全部移出确认",
            f"确定要将所有 {len(benched_names)} 个角色移出待选区吗？\n\n移出后这些角色将重新参与所有汇总统计。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            for name in benched_names:
                self.mgr.set_char_benched(name, False)
            if self.parent() and hasattr(self.parent(), "apply_filters"):
                self.parent().apply_filters()
                if hasattr(self.parent(), "statusBar"):
                    self.parent().statusBar().showMessage("已将所有角色移出待选区", 4000)
            self.load_bench_list()


def load_skill_descriptions():
    desc_path = os.path.join(os.path.dirname(__file__), "data", "bz_skill_desc.json")
    if os.path.exists(desc_path):
        try:
            with open(desc_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


class MainWindow(QMainWindow):
    def __init__(self, mgr):
        super().__init__()
        self.mgr = mgr
        self.all_chars = []
        self._current_cd_chars = []
        self.active_color_filter = None
        self.active_type_filter = None
        self._active_workers = set()
        self.skill_descs_map = load_skill_descriptions()

        self.setWindowTitle("剑网3 多角色周常管理助手 v1.0")
        self.setStyleSheet(DARK_QSS)
        icon = get_app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)

        self.init_ui()
        self.refresh_data()

    def _center_on_screen(self):
        screen = QApplication.primaryScreen()
        if screen:
            geo = screen.availableGeometry()
            x = (geo.width() - self.width()) // 2 + geo.x()
            y = (geo.height() - self.height()) // 2 + geo.y()
            self.move(x, y)

    def start_worker(self, thread):
        self._active_workers.add(thread)
        thread.finished.connect(lambda: self._on_worker_finished(thread))
        thread.start()

    def _on_worker_finished(self, thread):
        self._active_workers.discard(thread)
        thread.deleteLater()

    def closeEvent(self, event):
        try:
            settings = QSettings("JX3Manager", "JX3Manager")
            settings.setValue("geometry", self.saveGeometry())
            if hasattr(self, "tabs"):
                settings.setValue("current_tab", self.tabs.currentIndex())
        except Exception as e:
            logger.warning(f"保存窗口配置失败: {e}")

        for thread in list(self._active_workers):
            if thread.isRunning():
                thread.quit()
                thread.wait(1000)
            event.accept()

    def init_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        # Header Frame
        header = QFrame()
        header.setObjectName("HeaderFrame")
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(12, 8, 12, 8)
        
        lbl_title = QLabel("⚔ 剑网3 多角色周常管理器")
        lbl_title.setStyleSheet("font-size: 20px; font-weight: bold; color: #3b8ed0;")
        h_layout.addWidget(lbl_title)
        
        h_layout.addStretch()
        self.lbl_status = QLabel("就绪")
        self.lbl_status.setStyleSheet("color: #8888aa; font-size: 12px;")
        h_layout.addWidget(self.lbl_status)

        main_layout.addWidget(header)

        # Toolbar Frame (2-row clean layout)
        toolbar = QFrame()
        toolbar.setObjectName("ToolbarFrame")
        tb_main_layout = QVBoxLayout(toolbar)
        tb_main_layout.setContentsMargins(10, 8, 10, 8)
        tb_main_layout.setSpacing(8)

        # Row 1: Action Buttons
        tb_row1 = QHBoxLayout()
        tb_row1.setSpacing(10)

        self.btn_refresh = QPushButton("⟳  刷新本地数据")
        self.btn_refresh.setObjectName("PrimaryBtn")
        self.btn_refresh.clicked.connect(self.refresh_data)
        tb_row1.addWidget(self.btn_refresh)

        btn_stats_logs = QPushButton("⚙  一键开启全角色统计与战斗日志")
        btn_stats_logs.setObjectName("GreenBtn")
        btn_stats_logs.clicked.connect(self.configure_all_stats_and_logs)
        tb_row1.addWidget(btn_stats_logs)

        btn_json = QPushButton("⬇  JSON")
        btn_json.clicked.connect(self.export_json)
        tb_row1.addWidget(btn_json)

        btn_csv = QPushButton("⬇  CSV")
        btn_csv.clicked.connect(self.export_csv)
        tb_row1.addWidget(btn_csv)

        btn_bz = QPushButton("⚔️  百战招式概览")
        btn_bz.setObjectName("PrimaryBtn")
        btn_bz.setToolTip("点击打开独立的百战招式面板弹窗，查阅全角色百战精耐与招式细节")
        btn_bz.clicked.connect(lambda: self.open_baizhan_skills_dialog())
        tb_row1.addWidget(btn_bz)

        self.btn_api_config = QPushButton("🔑 API 设置")
        self.btn_api_config.setToolTip("查看或修改 JX3API Token（修改需二次确认）")
        self.btn_api_config.clicked.connect(self.open_api_config_dialog)
        tb_row1.addWidget(self.btn_api_config)

        tb_row1.addStretch()
        tb_main_layout.addLayout(tb_row1)

        # Row 2: Filter Controls (将筛选方式移动至下一行)
        tb_row2 = QHBoxLayout()
        tb_row2.setSpacing(10)

        # Server Filter
        tb_row2.addWidget(QLabel("服务器:"))
        self.combo_server = QComboBox()
        self.combo_server.addItem("所有服务器")
        self.combo_server.currentTextChanged.connect(self.apply_filters)
        tb_row2.addWidget(self.combo_server)

        tb_row2.addSpacing(10)

        # Search Filter
        tb_row2.addWidget(QLabel("搜索:"))
        self.input_search = QLineEdit()
        self.input_search.setPlaceholderText("角色/门派...")
        self.input_search.textChanged.connect(self.apply_filters)
        tb_row2.addWidget(self.input_search)

        tb_row2.addSpacing(15)

        # Equip Score Filter (Checkbox + SpinBox)
        self.chk_equip_score = QCheckBox("仅显示装分≥")
        self.chk_equip_score.setStyleSheet("color: #e0e0e0; font-size: 12px;")
        self.chk_equip_score.stateChanged.connect(self.apply_filters)
        tb_row2.addWidget(self.chk_equip_score)

        self.spin_equip_score = QDoubleSpinBox()
        self.spin_equip_score.setRange(0, 100)
        self.spin_equip_score.setValue(50.0)
        self.spin_equip_score.setSingleStep(5.0)
        self.spin_equip_score.setDecimals(1)
        self.spin_equip_score.setSuffix(" 万")
        self.spin_equip_score.setFixedWidth(115)
        self.spin_equip_score.valueChanged.connect(self.apply_filters)
        tb_row2.addWidget(self.spin_equip_score)

        tb_row2.addSpacing(15)

        # Stale Only Filter (Checkbox)
        self.chk_stale_only = QCheckBox("仅看本周未上线")
        self.chk_stale_only.setStyleSheet("color: #e0e0e0; font-size: 12px;")
        self.chk_stale_only.stateChanged.connect(self.apply_filters)
        tb_row2.addWidget(self.chk_stale_only)

        tb_row2.addSpacing(15)

        # Show Bench Filter (Checkbox)
        self.chk_show_bench = QCheckBox("显示待选区角色")
        self.chk_show_bench.setStyleSheet("color: #e0e0e0; font-size: 12px;")
        self.chk_show_bench.setToolTip("勾选后在表格中一并显示待选区角色（灰色斜体标记），但它们仍不计入汇总统计")
        self.chk_show_bench.setChecked(False)
        self.chk_show_bench.stateChanged.connect(self.apply_filters)
        tb_row2.addWidget(self.chk_show_bench)

        self.btn_bench_count = QPushButton("🪑 待选区 0 人")
        self.btn_bench_count.setObjectName("LinkBtn")
        self.btn_bench_count.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: none;
                color: #8888aa;
                font-size: 12px;
                padding: 2px 6px;
                text-decoration: underline;
                cursor: pointer;
            }
            QPushButton:hover {
                color: #b0bec5;
            }
        """)
        self.btn_bench_count.setToolTip("点击打开待选区管理面板")
        self.btn_bench_count.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_bench_count.setVisible(False)
        self.btn_bench_count.clicked.connect(self.open_bench_manager_dialog)
        tb_row2.addWidget(self.btn_bench_count)

        tb_row2.addStretch()
        tb_main_layout.addLayout(tb_row2)

        main_layout.addWidget(toolbar)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self.on_main_tab_changed)
        
        # Tab 1: Role Stats
        self.tab_roles = QWidget()
        layout_r = QVBoxLayout(self.tab_roles)
        self.table_roles = QTableWidget()
        self.setup_roles_table()
        layout_r.addWidget(self.table_roles)
        self.tabs.addTab(self.tab_roles, "角色统计概览")

        # Tab 2: Dungeon CD
        self.tab_cd = QWidget()
        layout_cd = QVBoxLayout(self.tab_cd)
        
        # Header bar for Dungeon CD tab
        cd_bar = QHBoxLayout()
        self.lbl_cd_status = QLabel("📅 武林通鉴周常活动日历")
        self.lbl_cd_status.setStyleSheet("font-size: 13px; font-weight: bold; color: #3b8ed0;")
        cd_bar.addWidget(self.lbl_cd_status)

        cd_bar.addSpacing(15)
        self.chk_show_legacy_cd = QCheckBox("显示过气副本")
        self.chk_show_legacy_cd.setChecked(False)
        self.chk_show_legacy_cd.setStyleSheet("color: #e0e0e0; font-size: 12px;")
        self.chk_show_legacy_cd.stateChanged.connect(self.apply_filters)
        cd_bar.addWidget(self.chk_show_legacy_cd)

        cd_bar.addStretch()

        self.btn_refresh_cal = QPushButton("⟳ 强制在线刷新活动日历")
        self.btn_refresh_cal.setStyleSheet("""
            QPushButton {
                background-color: #0d47a1; color: white; font-weight: bold;
                border-radius: 4px; padding: 4px 10px;
            }
            QPushButton:hover { background-color: #1565c0; }
        """)
        self.btn_refresh_cal.clicked.connect(self.refresh_active_calendar)
        cd_bar.addWidget(self.btn_refresh_cal)

        layout_cd.addLayout(cd_bar)

        # Summary bar for Dungeon CD tab
        cd_summary_bar = QHBoxLayout()
        self.lbl_cd_summary = QLabel()
        self.lbl_cd_summary.setStyleSheet("color: #b0bec5; font-size: 12px; font-weight: bold;")
        cd_summary_bar.addWidget(self.lbl_cd_summary)
        cd_summary_bar.addStretch()

        self.btn_uncleared_list = QPushButton("⚔ 查看未全清名单")
        self.btn_uncleared_list.setStyleSheet("""
            QPushButton {
                background-color: #37474f; color: #eceff1; font-size: 11px; font-weight: bold;
                border: 1px solid #546e7a; border-radius: 4px; padding: 4px 10px;
            }
            QPushButton:hover { background-color: #455a64; }
        """)
        self.btn_uncleared_list.setToolTip("点击查看当前未全清百战角色的详细名单")
        self.btn_uncleared_list.clicked.connect(self.show_uncleared_list_dialog)
        cd_summary_bar.addWidget(self.btn_uncleared_list)

        layout_cd.addLayout(cd_summary_bar)

        self.table_cd = QTableWidget()
        self.setup_cd_table()
        layout_cd.addWidget(self.table_cd)
        self.tabs.addTab(self.tab_cd, "副本 CD 统计")

        # Tab 3: Baizhan Weekly Boss Roster
        self.tab_roster = QWidget()
        self.setup_roster_tab()
        self.tabs.addTab(self.tab_roster, "本周百战BOSS排班")

        main_layout.addWidget(self.tabs)

        # 恢复窗口状态（位置、尺寸及当前 Tab）
        settings = QSettings("JX3Manager", "JX3Manager")
        geom = settings.value("geometry")
        restored = False
        if geom:
            restored = self.restoreGeometry(geom)
        if not restored:
            self.resize(1280, 800)
            self._center_on_screen()

        saved_tab = settings.value("current_tab", 0, type=int)
        if isinstance(saved_tab, int) and 0 <= saved_tab < self.tabs.count():
            self.tabs.setCurrentIndex(saved_tab)

    def setup_roles_table(self):
        headers = ["角色", "区服", "门派", "等级", "装备分", "资历", "金币", "休闲点", "侠义/威望", "百战精耐", "本地数据时间", "备注", "备注\n（每周重置）"]
        self.table_roles.setColumnCount(len(headers))
        self.table_roles.setHorizontalHeaderLabels(headers)
        self.table_roles.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_roles.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_roles.verticalHeader().setDefaultSectionSize(30)
        self.table_roles.setStyleSheet("QTableWidget { font-size: 12px; }")
        self.table_roles.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_roles.setAlternatingRowColors(True)
        self.table_roles.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_roles.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table_roles.setSortingEnabled(True)
        self.table_roles.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_roles.customContextMenuRequested.connect(lambda pos: self.show_char_context_menu(self.table_roles, pos))
        self.table_roles.itemDoubleClicked.connect(self.on_role_double_clicked)
        self.table_roles.itemSelectionChanged.connect(self.on_roles_selection_changed)

    def setup_cd_table(self):
        headers = ["角色", "区服", "百战进度\n(0-100)", "百战修罗", "换将点数", "本地数据时间", "备注", "备注\n（每周重置）"]
        self.table_cd.setColumnCount(len(headers))
        self.table_cd.setHorizontalHeaderLabels(headers)
        self.table_cd.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_cd.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_cd.verticalHeader().setDefaultSectionSize(30)
        self.table_cd.setStyleSheet("QTableWidget { font-size: 12px; }")
        
        header = self.table_cd.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.table_cd.setColumnWidth(0, 110)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.table_cd.setColumnWidth(1, 130)
        for col in range(2, len(headers)):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)

        self.table_cd.setAlternatingRowColors(True)
        self.table_cd.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_cd.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table_cd.setSortingEnabled(True)
        self.table_cd.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_cd.customContextMenuRequested.connect(lambda pos: self.show_char_context_menu(self.table_cd, pos))
        self.table_cd.itemDoubleClicked.connect(self.on_role_double_clicked)
        self.table_cd.itemSelectionChanged.connect(self.on_cd_selection_changed)

    def show_char_context_menu(self, table, pos):
        item_at_pos = table.itemAt(pos)
        if item_at_pos:
            r = item_at_pos.row()
            first_it = table.item(r, 0)
            if first_it and not first_it.isSelected():
                table.clearSelection()
                table.selectRow(r)

        selected_names = list(self.get_table_selected_char_names(table))
        if not selected_names:
            return

        bench_mgr = getattr(self.mgr, "bench_mgr", None)
        benched_in_sel = [name for name in selected_names if bench_mgr and bench_mgr.is_benched(name)]
        active_in_sel = [name for name in selected_names if not (bench_mgr and bench_mgr.is_benched(name))]

        menu = QMenu(self)
        menu.setStyleSheet(DARK_QSS)

        act_add_bench = None
        act_rem_bench = None

        if active_in_sel:
            if len(active_in_sel) == 1:
                act_add_bench = menu.addAction("🪑 移入待选区（不参与统计）")
            else:
                act_add_bench = menu.addAction(f"🪑 将选中的 {len(active_in_sel)} 个角色移入待选区")

        if benched_in_sel:
            if len(benched_in_sel) == 1:
                act_rem_bench = menu.addAction("↩ 移出待选区")
            else:
                act_rem_bench = menu.addAction(f"↩ 将选中的 {len(benched_in_sel)} 个角色移出待选区")

        if not menu.actions():
            return

        action = menu.exec(table.viewport().mapToGlobal(pos))
        if action == act_add_bench and active_in_sel:
            names_str = "、".join(active_in_sel[:10]) + ("..." if len(active_in_sel) > 10 else "")
            reply = QMessageBox.question(
                self, "移入待选区确认",
                f"确定要将以下 {len(active_in_sel)} 个角色移入待选区吗？\n\n【{names_str}】\n\n移入后这些角色将不计入任何汇总统计。",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                for name in active_in_sel:
                    self.mgr.set_char_benched(name, True)
                self.apply_filters()
                self.statusBar().showMessage(f"已将 {len(active_in_sel)} 个角色移入待选区", 4000)

        elif action == act_rem_bench and benched_in_sel:
            for name in benched_in_sel:
                self.mgr.set_char_benched(name, False)
            self.apply_filters()
            self.statusBar().showMessage(f"已将 {len(benched_in_sel)} 个角色移出待选区", 4000)

    def open_bench_manager_dialog(self):
        dlg = BenchManagerDialog(self.mgr, parent=self)
        dlg.exec()

    def get_table_selected_char_names(self, table):
        selected_names = set()
        for item in table.selectedItems():
            r = item.row()
            it = table.item(r, 0)
            if it and it.text().strip():
                t = it.text().strip()
                if t.startswith("🪑 "):
                    t = t[2:].strip()
                selected_names.add(t)
        return selected_names

    def select_chars_in_table(self, table, target_names):
        if not target_names:
            return
        table.blockSignals(True)
        table.clearSelection()
        selection_model = table.selectionModel()
        if not selection_model:
            table.blockSignals(False)
            return

        from PyQt6.QtCore import QItemSelection, QItemSelectionModel
        selection = QItemSelection()
        
        for row in range(table.rowCount()):
            name_item = table.item(row, 0)
            if name_item:
                raw_t = name_item.text().strip()
                if raw_t.startswith("🪑 "):
                    raw_t = raw_t[2:].strip()
                if raw_t in target_names:
                    first_index = table.model().index(row, 0)
                    last_index = table.model().index(row, table.columnCount() - 1)
                    selection.select(first_index, last_index)

        selection_model.select(selection, QItemSelectionModel.SelectionFlag.Select)
        table.blockSignals(False)

    def on_roles_selection_changed(self):
        if getattr(self, "_is_syncing_selection", False) or getattr(self, "_updating_tables", False):
            return
        names = self.get_table_selected_char_names(self.table_roles)
        if names:
            self.last_selected_char_names = names

    def on_cd_selection_changed(self):
        if getattr(self, "_is_syncing_selection", False) or getattr(self, "_updating_tables", False):
            return
        names = self.get_table_selected_char_names(self.table_cd)
        if names:
            self.last_selected_char_names = names

    def on_main_tab_changed(self, index):
        target_names = getattr(self, "last_selected_char_names", set())
        if not target_names:
            return

        self._is_syncing_selection = True
        try:
            if index == 0:  # 角色统计概览
                self.select_chars_in_table(self.table_roles, target_names)
            elif index == 1:  # 副本 CD 统计
                self.select_chars_in_table(self.table_cd, target_names)
        finally:
            self._is_syncing_selection = False

    def on_role_double_clicked(self, item):
        table = item.tableWidget()
        row = item.row()
        col = item.column()
        name_item = table.item(row, 0)
        if not name_item:
            return
        char_name = name_item.text().strip()
        if char_name.startswith("🪑 "):
            char_name = char_name[2:].strip()

        header_item = table.horizontalHeaderItem(col)
        header_text = header_item.text() if header_item else ""

        if table == self.table_roles:
            if col in (11, 12):
                is_weekly = (col == 12)
                note_type_str = "每周重置备注" if is_weekly else "常驻备注"
                self.edit_char_note(char_name, is_weekly, note_type_str)
                return
            if col == 9:
                self.open_baizhan_skills_dialog(char_name)
                return
        elif table == self.table_cd:
            if "换将点数" in header_text:
                self.edit_huanjiang_points(char_name, item)
                return
            if "每周重置" in header_text or col == table.columnCount() - 1:
                self.edit_char_note(char_name, True, "每周重置备注")
                return
            if "备注" in header_text or col == table.columnCount() - 2:
                self.edit_char_note(char_name, False, "常驻备注")
                return

        # Double-clicking character name / row opens RoleDetailDialog
        c = self.mgr.characters.get(char_name)
        if c:
            dlg = RoleDetailDialog(c, self.mgr.my_data, parent=self)
            dlg.exec()

    def edit_char_note(self, char_name, is_weekly, note_type_str):
        notes_mgr = getattr(self.mgr, "notes_mgr", None)
        if not notes_mgr:
            return

        old_perm, old_weekly = notes_mgr.get_note(char_name)
        old_val = old_weekly if is_weekly else old_perm

        dlg = StickyNoteDialog(char_name, note_type_str, initial_text=old_val, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_val = dlg.result_text
            if is_weekly:
                notes_mgr.set_weekly_note(char_name, new_val)
            else:
                notes_mgr.set_perm_note(char_name, new_val)

            # Sync across both tables
            self.sync_note_to_roles_table(char_name, is_weekly, new_val)
            self.sync_note_to_cd_table(char_name, is_weekly, new_val)
            self.lbl_status.setText(f"✓ 已成功保存角色【{char_name}】的{note_type_str}")
            self.lbl_status.setStyleSheet("color: #4caf50; font-weight: bold;")

    def open_baizhan_skills_dialog(self, default_char_name=None):
        active_chars = filter_out_benched(self.all_chars)
        dlg = BaizhanSkillsDialog(self.mgr, active_chars, default_char_name=default_char_name, parent=self)
        dlg.exec()

    def edit_huanjiang_points(self, char_name, item):
        try:
            curr_val = int(item.text()) if item.text().isdigit() else 0
        except Exception:
            curr_val = 0
            
        new_val, ok = QInputDialog.getInt(
            self, "修改换将点数",
            f"请输入角色【{char_name}】的换将点数:",
            curr_val, 0, 999999, 1
        )
        if ok and new_val != curr_val:
            reply = QMessageBox.question(
                self, "二次确认 - 修改换将点数",
                f"⚠️ 请确认是否要将角色【{char_name}】的换将点数修改为 {new_val}？\n\n（原点数: {curr_val} ➜ 新点数: {new_val}）",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.mgr.update_huanjiang_points(char_name, new_val)
                item.setText(str(new_val))
                if hasattr(item, "val"):
                    item.val = new_val

                new_up = None
                if char_name in self.mgr.characters:
                    new_up = self.mgr.characters[char_name].get("huanjiang_updated_at")
                up_str = new_up if new_up else "未知（旧数据，修改一次后自动记录）"
                item.setToolTip(f"💡 双击此单元格可修改换将点数（含防误触二次确认）\n🕐 数据填入时间：{up_str}")

                for c in self.all_chars:
                    if c.get("name") == char_name:
                        c["huanjiang_points"] = new_val
                        c["huanjiang_updated_at"] = new_up
                        break
                self._update_cd_summary(getattr(self, "_current_cd_chars", self.all_chars))
                self.lbl_status.setText(f"✓ 已成功将角色【{char_name}】的换将点数修改为 {new_val}")

    def show_uncleared_list_dialog(self):
        chars = getattr(self, "_current_cd_chars", self.all_chars)
        uncleared = []
        for c in chars:
            prog = c.get("baizhan_progress")
            killed = prog.get("killed", 0) if prog else 0
            total = prog.get("total", 0) if prog else 0
            is_cleared = (prog is not None) and (killed >= BAIZHAN_CLEARED_THRESHOLD or (killed >= total > 0))
            if not is_cleared:
                cname = c.get("name", "未知")
                prog_str = f"{killed}/{total}" if (prog and total > 0) else (str(killed) if prog else "-")
                stale_str = "（未上线）" if c.get("is_stale", False) else ""
                uncleared.append(f"{cname} {prog_str}{stale_str}")

        if not uncleared:
            QMessageBox.information(self, "百战未全清名单", "当前展示的所有角色百战均已全清！")
            return

        max_rows = 40
        if len(uncleared) > max_rows:
            lines = uncleared[:max_rows]
            lines.append(f"…等 {len(uncleared)} 个")
        else:
            lines = uncleared

        msg_text = "\n".join(lines)
        QMessageBox.information(self, "百战未全清角色名单", msg_text)

    def _update_cd_summary(self, chars, hidden_count=None):
        total_chars = len(chars)
        online_count = sum(1 for c in chars if not c.get("is_stale", False) and (c.get("record_time") or 0) > 0)
        stale_count = sum(1 for c in chars if c.get("is_stale", False))

        bz_chars = [c for c in chars if c.get("baizhan_progress")]
        bz_total = len(bz_chars)
        bz_cleared = sum(
            1 for c in bz_chars
            if c["baizhan_progress"].get("killed", 0) >= BAIZHAN_CLEARED_THRESHOLD
            or c["baizhan_progress"].get("killed", 0) >= c["baizhan_progress"].get("total", 0) > 0
        )

        xiuluo_count = sum(1 for c in chars if (c.get("baizhan_progress") or {}).get("xiuluo", False))
        hj_total = sum(int(c.get("huanjiang_points", 0) or 0) for c in chars)

        if hidden_count is None:
            hidden_count = getattr(self, "_hidden_cd_count", 0)

        summary_text = (
            f"👥 角色总数 {total_chars} | "
            f"🟢 本周已上线 {online_count} | "
            f"🔴 未上线 {stale_count} | "
            f"⚔ 百战全清 {bz_cleared}/{bz_total} | "
            f"👹 修罗完成 {xiuluo_count} | "
            f"🎯 换将点数总计 {hj_total}"
        )
        if hidden_count > 0:
            summary_text += f" | 📦 隐藏过气副本列 {hidden_count}"

        if hasattr(self, "lbl_cd_summary"):
            self.lbl_cd_summary.setText(summary_text)

    def on_edit_xiuluo_boss_clicked(self):
        wb = getattr(self.mgr, "weekly_bosses", {}) or {}
        default_b = wb.get("boss", "") if isinstance(wb, dict) else ""
        curr_b = getattr(self.mgr, "custom_xiuluo_boss", None) or default_b
        
        text, ok = QInputDialog.getText(
            self, "校核修正 - 本周修罗/镇守Boss",
            "如果API返回数据有误，请输入正确的本周修罗/镇守Boss名称:\n(多个Boss可用斜杠 / 分隔；若清空输入将重置为API默认值)",
            QLineEdit.EchoMode.Normal,
            curr_b
        )
        if ok:
            new_boss = text.strip()
            self.mgr.update_custom_xiuluo_boss(new_boss)
            self.update_roster_table()
            if hasattr(self, "on_data_loaded") and self.mgr.characters:
                self.on_data_loaded(self.mgr.characters)
            if new_boss:
                self.lbl_status.setText(f"✓ 已成功校核修正本周修罗Boss为: 【{new_boss}】，全员进度已重新校对！")
            else:
                self.lbl_status.setText("✓ 已重置本周修罗Boss为 API 接口默认值。")

    def configure_all_stats_and_logs(self):
        try:
            r1 = enable_all_stats()
            r2 = enable_combat_logs_for_all()
            if r1 and r2:
                QMessageBox.information(self, "配置完成", "✓ 已成功为所有角色开启【开启数据统计】与【保存战斗日志】功能！")
            else:
                QMessageBox.warning(self, "配置完成", f"配置部分完成：开启统计 ({'成功' if r1 else '失败'}), 开启日志 ({'成功' if r2 else '失败'})")
        except Exception as e:
            QMessageBox.critical(self, "配置异常", f"开启功能时发生异常: {e}")

    def sync_note_to_cd_table(self, char_name, is_weekly, text):
        self._updating_tables = True
        try:
            col = self.table_cd.columnCount() - (1 if is_weekly else 2)
            if col < 0:
                return
            for r in range(self.table_cd.rowCount()):
                it_n = self.table_cd.item(r, 0)
                if it_n and (it_n.text() == char_name or it_n.text() == f"🪑 {char_name}"):
                    target_item = self.table_cd.item(r, col)
                    if target_item and target_item.text() != text:
                        target_item.setText(text)
                    break
        finally:
            self._updating_tables = False

    def sync_note_to_roles_table(self, char_name, is_weekly, text):
        self._updating_tables = True
        try:
            col = 12 if is_weekly else 11
            for r in range(self.table_roles.rowCount()):
                it_n = self.table_roles.item(r, 0)
                if it_n and (it_n.text() == char_name or it_n.text() == f"🪑 {char_name}"):
                    target_item = self.table_roles.item(r, col)
                    if target_item and target_item.text() != text:
                        target_item.setText(text)
                    break
        finally:
            self._updating_tables = False

    def refresh_data(self):
        if hasattr(self, "btn_refresh"):
            self.btn_refresh.setEnabled(False)
        self.lbl_status.setText("⌛ 正在加载本地角色数据...")
        self.thread = DataLoaderThread(self.mgr)
        self.thread.loaded.connect(self.on_data_loaded)
        self.thread.error.connect(self._on_data_load_error)
        self.start_worker(self.thread)

    def _on_data_load_error(self, err):
        if hasattr(self, "btn_refresh"):
            self.btn_refresh.setEnabled(True)
        self.lbl_status.setText(f"❌ 加载失败: {err}")

    def on_data_loaded(self, chars):
        if hasattr(self, "btn_refresh"):
            self.btn_refresh.setEnabled(True)
        self.all_chars = list(chars.values())
        
        # Populate server list
        servers = sorted(list({f"{c.get('region', '')}/{c.get('server', '')}".strip("/") for c in self.all_chars if c.get("server")}))
        self.combo_server.blockSignals(True)
        self.combo_server.clear()
        self.combo_server.addItem("所有服务器")
        self.combo_server.addItems(servers)
        self.combo_server.blockSignals(False)

        self.apply_filters()
        self.update_roster_table()
        
        from readers.baizhan_api import api as bz_api
        if bz_api.is_cache_stale(getattr(self.mgr, "weekly_bosses", {})):
            self.lbl_status.setText(f"⚠️ 已加载 {len(self.all_chars)} 个角色 [警告: 当前百战排班缓存已跨周陈旧，修罗/击杀进度建议在线刷新]")
            self.lbl_status.setStyleSheet("color: #ffa726; font-weight: bold;")
        else:
            self.lbl_status.setText(f"✓ 已加载 {len(self.all_chars)} 个角色 ({datetime.datetime.now().strftime('%H:%M:%S')})")
            self.lbl_status.setStyleSheet("")

    def apply_filters(self):
        selected_server = self.combo_server.currentText()
        search_kw = self.input_search.text().lower().strip()
        min_equip_w = self.spin_equip_score.value() if self.chk_equip_score.isChecked() else 0
        stale_only = self.chk_stale_only.isChecked() if hasattr(self, "chk_stale_only") else False
        show_bench = self.chk_show_bench.isChecked() if hasattr(self, "chk_show_bench") else False

        bench_mgr = getattr(self.mgr, "bench_mgr", None)
        bench_count = bench_mgr.count() if bench_mgr else sum(1 for c in self.all_chars if c.get("is_benched", False))
        if hasattr(self, "btn_bench_count"):
            if bench_count > 0:
                self.btn_bench_count.setText(f"🪑 待选区 {bench_count} 人")
                self.btn_bench_count.setVisible(True)
            else:
                self.btn_bench_count.setVisible(False)

        filtered_for_display = []
        filtered_for_stats = []

        for c in self.all_chars:
            is_benched = c.get("is_benched", False)
            if stale_only and not c.get("is_stale", False):
                continue
            srv = f"{c.get('region', '')}/{c.get('server', '')}".strip("/")
            if selected_server != "所有服务器" and srv != selected_server:
                continue
            name = c.get("name", "").lower()
            force = c.get("force_name", "").lower()
            if search_kw and (search_kw not in name and search_kw not in force):
                continue

            if min_equip_w > 0:
                score = int(c.get("equip_score", 0))
                if score < min_equip_w * 10000:
                    continue

            if not is_benched:
                filtered_for_stats.append(c)
                filtered_for_display.append(c)
            elif show_bench:
                filtered_for_display.append(c)

        self.update_roles_table(filtered_for_display)
        self.update_cd_table(filtered_for_display, stats_chars=filtered_for_stats)

        target_names = getattr(self, "last_selected_char_names", set())
        if target_names:
            curr_tab = self.tabs.currentIndex()
            if curr_tab == 0:
                self.select_chars_in_table(self.table_roles, target_names)
            elif curr_tab == 1:
                self.select_chars_in_table(self.table_cd, target_names)

    def update_roles_table(self, chars):
        self._updating_tables = True
        try:
            self.table_roles.setSortingEnabled(False)
            self.table_roles.setRowCount(len(chars))
            notes_mgr = getattr(self.mgr, "notes_mgr", None)
            for i, c in enumerate(chars):
                cname = c.get("name", "")
                is_benched = c.get("is_benched", False)
                disp_name = f"🪑 {cname}" if is_benched else cname
                bz_api = c.get("baizhan_api", {}) or {}
                stamina = bz_api.get("skillStamina")
                energy = bz_api.get("skillEnergy")
                if stamina is not None and energy is not None:
                    bz_jn_str = f"{stamina}/{energy}"
                    bz_jn_val = stamina + energy
                else:
                    bz_jn_str = "-"
                    bz_jn_val = -1
                srv = f"{c.get('region', '')}/{c.get('server', '')}".strip("/")

                perm_n, weekly_n = notes_mgr.get_note(cname) if notes_mgr else ("", "")
                it_p = QTableWidgetItem(perm_n)
                it_w = QTableWidgetItem(weekly_n)
                it_p.setToolTip("💡 双击可编辑常驻备注（保存时含二次确认防误触）")
                it_w.setToolTip("💡 双击可编辑每周重置备注（每周一 12:00 自动重置，保存时含二次确认防误触）")

                items = [
                    QTableWidgetItem(disp_name),
                    QTableWidgetItem(srv),
                    QTableWidgetItem(c.get("force_name", "")),
                    NumericTableWidgetItem(c.get("level", 0), int(c.get("level", 0))),
                    NumericTableWidgetItem(c.get("equip_score", 0), int(c.get("equip_score", 0))),
                    NumericTableWidgetItem(c.get("achievement_score", 0), int(c.get("achievement_score", 0))),
                    NumericTableWidgetItem(c.get("gold", 0), int(c.get("gold", 0))),
                    NumericTableWidgetItem(c.get("contribution", 0), int(c.get("contribution", 0))),
                    NumericTableWidgetItem(c.get("justice", 0), int(c.get("justice", 0))),
                    NumericTableWidgetItem(bz_jn_str, bz_jn_val),
                    QTableWidgetItem(c.get("last_update", "未知")),
                    it_p,
                    it_w
                ]
                for j, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    if is_benched:
                        item.setForeground(QColor("#777788"))
                        f = item.font()
                        f.setItalic(True)
                        item.setFont(f)
                    self.table_roles.setItem(i, j, item)
        finally:
            self.table_roles.setSortingEnabled(True)
            self._updating_tables = False

    def refresh_active_calendar(self):
        self.btn_refresh_cal.setEnabled(False)
        self.btn_refresh_cal.setText("⟳ 正在同步日历...")
        self.statusBar().showMessage("正在向 JX3API 强制刷新武林通鉴活动日历...")
        self.cal_thread = CalendarFetchThread(self.mgr)
        self.cal_thread.fetched.connect(self.on_calendar_fetched)
        self.start_worker(self.cal_thread)

    def on_calendar_fetched(self, cal):
        self.btn_refresh_cal.setEnabled(True)
        self.btn_refresh_cal.setText("⟳ 强制在线刷新活动日历")
        self.statusBar().showMessage("武林通鉴活动日历同步完成！", 4000)
        self.apply_filters()

    def update_cd_table(self, chars, stats_chars=None):
        if stats_chars is None:
            stats_chars = [c for c in chars if not c.get("is_benched", False)]
        self._current_cd_chars = stats_chars
        self._updating_tables = True
        try:
            self.table_cd.setSortingEnabled(False)

            calendar = getattr(self.mgr, "active_calendar", None)
            if not calendar:
                from readers.baizhan_api import api as bz_api
                calendar = bz_api.get_active_calendar(force_refresh=False) or {}

            if isinstance(calendar, dict) and "data" in calendar and isinstance(calendar["data"], dict):
                c_data = calendar["data"]
            else:
                c_data = calendar if isinstance(calendar, dict) else {}

            api_date = c_data.get("date", "未知") if isinstance(c_data, dict) else "未知"
            sync_time = c_data.get("_sync_time", "未知") if isinstance(c_data, dict) else "未知"
            if hasattr(self, "lbl_cd_status"):
                self.lbl_cd_status.setText(f"📅 武林通鉴周常日历  [API服务器数据日期: {api_date} | 客户端同步时间: {sync_time}]")

            # 获取本周 raid 名单
            weekly = c_data.get("weekly", {}) if isinstance(c_data, dict) else {}
            raid_names = weekly.get("raid", []) if isinstance(weekly, dict) else []
            if not isinstance(raid_names, list):
                raid_names = []

            # 收集本批 chars 所有出现过的副本 ID（按 ID 数值升序排序，排除百战562）
            from readers.dungeon_cd import DUNGEON_NAMES
            all_dids = set()
            for c in chars:
                dungeons = c.get("dungeon_cd", {})
                if isinstance(dungeons, dict):
                    for did in dungeons.keys():
                        try:
                            id_num = int(did)
                            if id_num != 562:
                                all_dids.add(id_num)
                        except (ValueError, TypeError):
                            pass
            sorted_dids = sorted(list(all_dids))

            # 过滤副本列
            show_legacy = self.chk_show_legacy_cd.isChecked() if hasattr(self, "chk_show_legacy_cd") else False
            visible_dids, hidden_dids = filter_cd_dungeon_ids(sorted_dids, DUNGEON_NAMES, raid_names, show_legacy=show_legacy)

            self._hidden_cd_count = len(hidden_dids)
            self._update_cd_summary(stats_chars, hidden_count=self._hidden_cd_count)

            dungeon_headers = [DUNGEON_NAMES.get(did, f"副本{did}") for did in visible_dids]

            headers = [
                "角色", "区服",
                *dungeon_headers,
                "百战进度\n(0-100)",
                "百战修罗",
                "换将点数",
                "本地数据时间",
                "备注",
                "备注\n（每周重置）"
            ]

            self.table_cd.setColumnCount(len(headers))
            self.table_cd.setHorizontalHeaderLabels(headers)
            self.table_cd.setRowCount(len(chars))

            # 设置列宽与自适应模式
            dungeon_col_indices = set(range(2, 2 + len(visible_dids)))
            header = self.table_cd.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
            self.table_cd.setColumnWidth(0, 110)
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
            self.table_cd.setColumnWidth(1, 130)

            for didx in dungeon_col_indices:
                header.setSectionResizeMode(didx, QHeaderView.ResizeMode.Interactive)
                self.table_cd.setColumnWidth(didx, 130)

            for rest_col in range(2 + len(visible_dids), len(headers)):
                header.setSectionResizeMode(rest_col, QHeaderView.ResizeMode.Stretch)

            notes_mgr = getattr(self.mgr, "notes_mgr", None)

            for i, c in enumerate(chars):
                srv = f"{c.get('region', '')}/{c.get('server', '')}".strip("/")
                name = c.get("name", "")
                is_benched = c.get("is_benched", False)
                disp_name = f"🪑 {name}" if is_benched else name
                dungeons = c.get("dungeon_cd", {})
                prog = c.get("baizhan_progress", {})
                is_stale = c.get("is_stale", False)

                name_item = QTableWidgetItem(disp_name)
                
                # 角色名 tooltip
                tt_lines = []
                if is_stale:
                    tt_lines.append("⚠️ 该角色本周一 12:00 之后尚未登录上线，副本 CD 已由服务端自动重置为 0。")

                # 检查未显示副本中的进度
                if hidden_dids and isinstance(dungeons, dict):
                    hidden_progs = []
                    for hdid in hidden_dids:
                        ddata = dungeons.get(hdid)
                        if ddata is None:
                            ddata = dungeons.get(str(hdid))
                        if isinstance(ddata, dict) and ddata:
                            done = ddata.get("done", ddata.get("doneCount", 0))
                            total = ddata.get("total", ddata.get("totalCount", 0))
                            hname = DUNGEON_NAMES.get(hdid, f"副本{hdid}")
                            hidden_progs.append(f"{hname} {done}/{total}")
                    if hidden_progs:
                        tt_lines.append(f"📦 未显示副本: {', '.join(hidden_progs[:5])}")

                if tt_lines:
                    name_item.setToolTip("\n".join(tt_lines))

                items = [
                    name_item,
                    QTableWidgetItem(srv)
                ]

                # 动态副本列 (仅可见列)
                for did in visible_dids:
                    ddata = dungeons.get(did)
                    if ddata is None:
                        ddata = dungeons.get(str(did))

                    if isinstance(ddata, dict) and ddata:
                        done = ddata.get("done", ddata.get("doneCount", 0))
                        total = ddata.get("total", ddata.get("totalCount", 0))
                        d_item = NumericTableWidgetItem(f"{done}/{total}", done)
                        if done >= total and total > 0:
                            d_item.setForeground(QColor("#4caf50"))
                            f = d_item.font()
                            f.setBold(True)
                            d_item.setFont(f)
                        elif 0 < done < total:
                            d_item.setForeground(QColor("#ffca28"))
                        elif done == 0 and total > 0:
                            d_item.setForeground(QColor("#ef5350"))
                        else:
                            d_item.setForeground(QColor("#888888"))
                        items.append(d_item)
                    else:
                        d_item = NumericTableWidgetItem("-", -1)
                        d_item.setForeground(QColor("#888888"))
                        items.append(d_item)

                # 百战进度列
                if prog:
                    killed = prog.get("killed", 0)
                    total = prog.get("total", 0)
                    bz_k_str = f"{killed}/{total}" if total > 0 else str(killed)
                    bz_k_val = killed
                    bz_item = NumericTableWidgetItem(bz_k_str, bz_k_val)
                    if (killed >= total and total > 0) or killed >= BAIZHAN_CLEARED_THRESHOLD:
                        bz_item.setForeground(QColor("#4caf50"))
                        f = bz_item.font()
                        f.setBold(True)
                        bz_item.setFont(f)
                    elif killed > 0:
                        bz_item.setForeground(QColor("#ffca28"))
                    elif killed == 0:
                        bz_item.setForeground(QColor("#ef5350"))
                    else:
                        bz_item.setForeground(QColor("#888888"))

                    bz_tt_lines = []
                    kb = prog.get("killed_bosses", [])
                    if kb:
                        bz_tt_lines.append(f"已击杀首领: {', '.join(kb)}")
                    um = prog.get("unmatched", [])
                    if um:
                        bz_tt_lines.append(f"轮换表外击杀: {', '.join(um)}")
                    if not bz_tt_lines:
                        bz_tt_lines.append("本周暂无击杀记录")
                    bz_item.setToolTip("\n".join(bz_tt_lines))
                else:
                    bz_item = NumericTableWidgetItem("-", -1)
                    bz_item.setForeground(QColor("#888888"))
                items.append(bz_item)

                # 百战修罗列
                bz_x = ("是" if prog.get("xiuluo") else "否") if prog else "-"
                x_item = QTableWidgetItem(bz_x)
                if bz_x == "是":
                    x_item.setForeground(QColor("#4caf50"))
                    f = x_item.font()
                    f.setBold(True)
                    x_item.setFont(f)
                elif bz_x == "否":
                    x_item.setForeground(QColor("#888888"))
                items.append(x_item)
                
                # 换将点数
                hj_pts = c.get("huanjiang_points", 0)
                hj_up = c.get("huanjiang_updated_at")
                hj_item = NumericTableWidgetItem(str(hj_pts), hj_pts)
                up_str = hj_up if hj_up else "未知（旧数据，修改一次后自动记录）"
                hj_item.setToolTip(f"💡 双击此单元格可修改换将点数（含防误触二次确认）\n🕐 数据填入时间：{up_str}")
                items.append(hj_item)
                
                # 本地数据时间
                items.append(QTableWidgetItem(c.get("last_update", "未知")))

                # 备注
                perm_n, weekly_n = notes_mgr.get_note(name) if notes_mgr else ("", "")
                it_p = QTableWidgetItem(perm_n)
                it_w = QTableWidgetItem(weekly_n)
                it_p.setToolTip("💡 双击可编辑常驻备注（保存时含二次确认防误触）")
                it_w.setToolTip("💡 双击可编辑每周重置备注（每周一 12:00 自动重置，保存时含二次确认防误触）")
                items.append(it_p)
                items.append(it_w)

                # is_stale 整行变色
                if is_stale:
                    for item in items:
                        item.setForeground(QColor("#78909c"))

                if is_benched:
                    for item in items:
                        item.setForeground(QColor("#777788"))
                        f = item.font()
                        f.setItalic(True)
                        item.setFont(f)

                for j, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    if is_stale and j in dungeon_col_indices:
                        item.setToolTip("⚠️ 该角色本周一 12:00 之后尚未登录上线，副本 CD 已自动重置。")
                    self.table_cd.setItem(i, j, item)
        finally:
            self.table_cd.setSortingEnabled(True)
            self._updating_tables = False

    def setup_roster_tab(self):
        layout = QVBoxLayout(self.tab_roster)
        
        # Banner bar
        banner = QFrame()
        banner.setStyleSheet("background-color: #1a1a3a; border-radius: 6px; padding: 6px 12px;")
        b_layout = QHBoxLayout(banner)
        b_layout.setContentsMargins(8, 6, 8, 6)

        self.lbl_roster_week = QLabel("⚔️ 本周百战异闻录 (第 ? 周)")
        self.lbl_roster_week.setStyleSheet("font-size: 13px; font-weight: bold; color: #3b8ed0;")
        b_layout.addWidget(self.lbl_roster_week)
        b_layout.addSpacing(20)

        self.lbl_roster_xiuluo = QLabel("👹 本周修罗/镇守首领: 【 - 】")
        self.lbl_roster_xiuluo.setStyleSheet("font-size: 14px; font-weight: bold; color: #ffca28;")
        b_layout.addWidget(self.lbl_roster_xiuluo)

        self.btn_edit_xiuluo = QPushButton("⚠️ API数据有错误")
        self.btn_edit_xiuluo.setToolTip("如果 API 返回的修罗/镇守 Boss 名称有误，点击手动纠正并校核全员百战修罗进度")
        self.btn_edit_xiuluo.setStyleSheet("""
            QPushButton {
                background-color: #d97706;
                color: #ffffff;
                font-size: 12px;
                font-weight: bold;
                border-radius: 4px;
                padding: 3px 8px;
                margin-left: 6px;
            }
            QPushButton:hover {
                background-color: #b45309;
            }
        """)
        self.btn_edit_xiuluo.clicked.connect(self.on_edit_xiuluo_boss_clicked)
        b_layout.addWidget(self.btn_edit_xiuluo)

        b_layout.addStretch()

        self.lbl_roster_dates = QLabel("📅 周期: -")
        self.lbl_roster_dates.setStyleSheet("font-size: 12px; color: #8a8ab0;")
        b_layout.addWidget(self.lbl_roster_dates)

        layout.addWidget(banner)

        # Stale Warning Banner Frame
        self.frame_roster_warning = QFrame()
        self.frame_roster_warning.setStyleSheet("background-color: #332a00; border: 1px solid #ffb300; border-radius: 6px; padding: 4px 10px;")
        w_layout = QHBoxLayout(self.frame_roster_warning)
        w_layout.setContentsMargins(8, 4, 8, 4)
        self.lbl_roster_warning = QLabel("⚠️ [警告] 当前百战排班为跨周离线旧缓存。击杀与修罗计算可能受到影响，建议点击右侧【⟳ 强制在线刷新排班】！")
        self.lbl_roster_warning.setStyleSheet("font-size: 12px; font-weight: bold; color: #ffe082;")
        w_layout.addWidget(self.lbl_roster_warning)
        self.frame_roster_warning.setVisible(False)
        layout.addWidget(self.frame_roster_warning)

        # Toolbar
        t_bar = QHBoxLayout()
        lbl_filter = QLabel("🔍 搜索首领/技能/效果:")
        self.input_roster_search = QLineEdit()
        self.input_roster_search.setPlaceholderText("搜索BOSS名、技能或效果...")
        self.input_roster_search.textChanged.connect(self.apply_roster_filters)
        t_bar.addWidget(lbl_filter)
        t_bar.addWidget(self.input_roster_search)

        lbl_floor = QLabel("层数区间:")
        self.combo_roster_floor = QComboBox()
        self.combo_roster_floor.addItems([
            "全部层数 (1-100层)",
            "1-10层 (初级)",
            "11-30层 (中级)",
            "31-50层 (高级)",
            "51-70层 (大师)",
            "71-90层 (宗师)",
            "91-100层 (巅峰)"
        ])
        self.combo_roster_floor.currentIndexChanged.connect(self.apply_roster_filters)
        t_bar.addWidget(lbl_floor)
        t_bar.addWidget(self.combo_roster_floor)

        self.chk_roster_effect = QCheckBox("仅看因陀罗/特殊效果层")
        self.chk_roster_effect.stateChanged.connect(self.apply_roster_filters)
        t_bar.addWidget(self.chk_roster_effect)

        t_bar.addSpacing(10)
        self.btn_refresh_roster = QPushButton("⟳ 强制在线刷新排班")
        self.btn_refresh_roster.setStyleSheet("""
            QPushButton {
                background-color: #0d47a1; color: white; font-weight: bold;
                border-radius: 4px; padding: 4px 10px;
            }
            QPushButton:hover { background-color: #1565c0; }
        """)
        self.btn_refresh_roster.clicked.connect(self.refresh_roster_data)
        t_bar.addWidget(self.btn_refresh_roster)

        # View Mode Switcher
        t_bar.addSpacing(15)
        self.btn_roster_table_view = QPushButton("📋 列表明细模式")
        self.btn_roster_snake_view = QPushButton("🧩 100关卡图谱 (图片同款)")
        
        self.btn_roster_table_view.setCheckable(True)
        self.btn_roster_snake_view.setCheckable(True)
        self.btn_roster_table_view.setChecked(True)

        self.btn_roster_table_view.setStyleSheet("""
            QPushButton { background-color: #222; color: #aaa; padding: 4px 10px; border-radius: 4px; font-weight: bold; }
            QPushButton:checked { background-color: #3b8ed0; color: white; }
        """)
        self.btn_roster_snake_view.setStyleSheet("""
            QPushButton { background-color: #222; color: #aaa; padding: 4px 10px; border-radius: 4px; font-weight: bold; }
            QPushButton:checked { background-color: #3b8ed0; color: white; }
        """)

        self.btn_roster_table_view.clicked.connect(lambda: self.switch_roster_view(0))
        self.btn_roster_snake_view.clicked.connect(lambda: self.switch_roster_view(1))

        t_bar.addWidget(self.btn_roster_table_view)
        t_bar.addWidget(self.btn_roster_snake_view)

        layout.addLayout(t_bar)

        # Stacked Container for Table View and Snake Grid View
        self.roster_stack = QStackedWidget()

        # Table View (Page 0)
        self.table_roster = QTableWidget()
        headers = ["层数", "首领名称", "首领主要招式", "因陀罗 / 特殊效果", "效果说明"]
        self.table_roster.setColumnCount(len(headers))
        self.table_roster.setHorizontalHeaderLabels(headers)
        self.table_roster.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table_roster.setColumnWidth(0, 70)
        self.table_roster.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table_roster.setColumnWidth(1, 140)
        self.table_roster.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table_roster.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.table_roster.setColumnWidth(3, 200)
        self.table_roster.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)

        self.table_roster.setAlternatingRowColors(True)
        self.table_roster.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_roster.setSortingEnabled(True)
        self.roster_stack.addWidget(self.table_roster)

        # Snake Grid Scroll View (Page 1)
        self.snake_scroll = QScrollArea()
        self.snake_scroll.setWidgetResizable(True)
        self.snake_scroll.setStyleSheet("QScrollArea { background-color: #f8fafc; border: 1px solid #cbd5e1; border-radius: 6px; }")

        self.snake_container = QWidget()
        self.snake_container.setStyleSheet("QWidget { background-color: #ffffff; }")
        self.snake_layout = QVBoxLayout(self.snake_container)
        self.snake_layout.setContentsMargins(15, 15, 15, 15)
        self.snake_layout.setSpacing(10)
        self.snake_scroll.setWidget(self.snake_container)

        self.roster_stack.addWidget(self.snake_scroll)

        layout.addWidget(self.roster_stack)

    def update_roster_table(self):
        if not hasattr(self, "table_roster"):
            return
        
        self.table_roster.setSortingEnabled(False)
        wb = getattr(self.mgr, "weekly_bosses", {}) or {}
        if isinstance(wb, dict):
            week_num = wb.get("week", "?")
            custom_b = getattr(self.mgr, "custom_xiuluo_boss", None)
            if custom_b:
                boss_name = f"{custom_b} (⚠️ 已校核修正)"
            else:
                boss_name = wb.get("boss", "未知")
            s_val = wb.get("start", "")
            e_val = wb.get("end", "")
            sync_time = wb.get("_sync_time", "未知")

            def format_date_item(v):
                if not v: return ""
                if isinstance(v, (int, float)) or (isinstance(v, str) and v.isdigit()):
                    try:
                        return datetime.datetime.fromtimestamp(int(v)).strftime("%Y-%m-%d")
                    except Exception: pass
                return str(v)

            s_str = format_date_item(s_val)
            e_str = format_date_item(e_val)
            api_date = f"{s_str} ~ {e_str}" if s_str and e_str else (s_str or "本周")
        else:
            week_num = "?"
            boss_name = "未知"
            api_date = "未知"
            sync_time = "未知"

        if hasattr(self, "lbl_roster_week"):
            self.lbl_roster_week.setText(f"⚔️ 本周百战异闻录 (第 {week_num} 周)")
        if hasattr(self, "lbl_roster_xiuluo"):
            self.lbl_roster_xiuluo.setText(f"👹 本周修罗/镇守首领: 【 {boss_name} 】")
        if hasattr(self, "lbl_roster_dates"):
            self.lbl_roster_dates.setText(f"📅 API周期日期: {api_date} | 客户端同步时间: {sync_time}")

        if hasattr(self, "lbl_roster_warning"):
            from readers.baizhan_api import api as bz_api
            is_stale = bz_api.is_cache_stale(wb)
            if is_stale:
                self.lbl_roster_warning.setText(
                    f"⚠️ [警告] 当前百战排班为跨周离线旧缓存(第 {week_num} 周)。击杀与修罗计算可能不准，建议点击右侧【⟳ 强制在线刷新排班】！"
                )
                self.frame_roster_warning.setVisible(True)
            else:
                self.frame_roster_warning.setVisible(False)

        boss_list = wb.get("list", []) if isinstance(wb, dict) and "list" in wb else (wb.get("data", []) if isinstance(wb, dict) else [])
        if isinstance(wb, list):
            boss_list = wb

        self.all_roster_data = boss_list
        self.apply_roster_filters()

    def switch_roster_view(self, idx):
        self.btn_roster_table_view.setChecked(idx == 0)
        self.btn_roster_snake_view.setChecked(idx == 1)
        self.roster_stack.setCurrentIndex(idx)
        if idx == 1:
            self.render_roster_snake_grid()

    def render_roster_snake_grid(self):
        while self.snake_layout.count():
            child = self.snake_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        wb = getattr(self.mgr, "weekly_bosses", {}) or {}
        boss_list = wb.get("list", []) if isinstance(wb, dict) and "list" in wb else []
        if not boss_list:
            lbl = QLabel("暂无本周百战 100 层排班数据。")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.snake_layout.addWidget(lbl)
            return

        # Header Title Banner (Matches 示例图片1.png)
        head_banner = QFrame()
        head_banner.setStyleSheet("background-color: #ffffff; border-bottom: 2px solid #e2e8f0; padding-bottom: 12px;")
        hb_layout = QVBoxLayout(head_banner)
        hb_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        lbl_t1 = QLabel("百 战 异 闻 录")
        lbl_t1.setStyleSheet("font-size: 22px; font-weight: 900; color: #1e293b; letter-spacing: 4px; border: none;")
        lbl_t1.setAlignment(Qt.AlignmentFlag.AlignCenter)

        week_num = wb.get("week", "?")
        custom_b = getattr(self.mgr, "custom_xiuluo_boss", None)
        if custom_b:
            boss_name = f"{custom_b} (⚠️ 已校核修正)"
        else:
            boss_name = wb.get("boss", "未知")
        s_val = wb.get("start", "")
        e_val = wb.get("end", "")

        def format_d(v):
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.isdigit()):
                try: return datetime.datetime.fromtimestamp(int(v)).strftime("%Y-%m-%d %H:%M")
                except: pass
            return str(v)

        s_str = format_d(s_val)
        e_str = format_d(e_val)
        d_str = f"{s_str} — {e_str}" if s_str and e_str else "本周"
        lbl_sub = QLabel(f"第 {week_num} 周  ·  {boss_name}  ·  {d_str}")
        lbl_sub.setStyleSheet("font-size: 13px; color: #64748b; font-weight: bold; border: none;")
        lbl_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)

        hb_layout.addWidget(lbl_t1)
        hb_layout.addWidget(lbl_sub)
        self.snake_layout.addWidget(head_banner)

        # 10 Rows Snake Grid
        boss_map = {b.get("index", i+1): b for i, b in enumerate(boss_list)}

        grid_frame = QWidget()
        grid_vbox = QVBoxLayout(grid_frame)
        grid_vbox.setSpacing(10)
        grid_vbox.setContentsMargins(5, 5, 5, 5)

        for row in range(10):
            row_box = QHBoxLayout()
            row_box.setSpacing(8)
            row_box.setAlignment(Qt.AlignmentFlag.AlignCenter)

            if row % 2 == 0:
                # Left to Right
                floors = [row * 10 + col + 1 for col in range(10)]
            else:
                # Right to Left
                floors = [(row + 1) * 10 - col for col in range(10)]

            for f_idx in floors:
                f_data = boss_map.get(f_idx, {"index": f_idx, "name": "未知", "data": {"name": "无"}})
                node = RosterNodeWidget(f_data)
                row_box.addWidget(node)

            grid_vbox.addLayout(row_box)

        self.snake_layout.addWidget(grid_frame)

    def refresh_roster_data(self):
        # 刷新排班时自动清除用户自定义的修罗首领名称
        if getattr(self.mgr, "custom_xiuluo_boss", None):
            self.mgr.update_custom_xiuluo_boss("")
        if hasattr(self, "btn_refresh_roster"):
            self.btn_refresh_roster.setEnabled(False)
            self.btn_refresh_roster.setText("⟳ 正在同步排班...")
        self.statusBar().showMessage("正在向 JX3API 强制刷新本周百战首领排班并重置修罗校核记录...")
        self.roster_thread = RosterFetchThread(self.mgr)
        self.roster_thread.fetched.connect(self.on_roster_fetched)
        self.start_worker(self.roster_thread)

    def on_roster_fetched(self, wb):
        if hasattr(self, "btn_refresh_roster"):
            self.btn_refresh_roster.setEnabled(True)
            self.btn_refresh_roster.setText("⟳ 强制在线刷新排班")
        self.statusBar().showMessage("百战首领排班同步完成，已重置修罗Boss校核！", 4000)
        self.update_roster_table()
        if hasattr(self, "on_data_loaded") and self.mgr.characters:
            self.on_data_loaded(self.mgr.characters)

    def apply_roster_filters(self):
        if not hasattr(self, "all_roster_data"):
            return
        
        self.table_roster.setSortingEnabled(False)
        q = self.input_roster_search.text().strip().lower()
        floor_idx = self.combo_roster_floor.currentIndex()
        only_effect = self.chk_roster_effect.isChecked()

        filtered = []
        for item in self.all_roster_data:
            idx = item.get("index", 0)
            name = item.get("name", "")
            skills = " / ".join(item.get("skill", []))
            eff_data = item.get("data", {}) if isinstance(item.get("data"), dict) else {}
            eff_name = eff_data.get("name", "无")
            eff_desc = eff_data.get("desc", "无")

            # Floor range filter
            if floor_idx == 1 and not (1 <= idx <= 10): continue
            elif floor_idx == 2 and not (11 <= idx <= 30): continue
            elif floor_idx == 3 and not (31 <= idx <= 50): continue
            elif floor_idx == 4 and not (51 <= idx <= 70): continue
            elif floor_idx == 5 and not (71 <= idx <= 90): continue
            elif floor_idx == 6 and not (91 <= idx <= 100): continue

            # Only effect filter
            has_effect = eff_name not in ("无", "", None) and eff_name != "无"
            if only_effect and not has_effect:
                continue

            # Text search query filter
            if q:
                match_text = f"{idx} {name} {skills} {eff_name} {eff_desc}".lower()
                if q not in match_text:
                    continue

            filtered.append(item)

        self.table_roster.setRowCount(len(filtered))
        for i, b in enumerate(filtered):
            idx = b.get("index", 0)
            name = b.get("name", "")
            skills = "  ".join([f"[{s}]" for s in b.get("skill", [])])
            eff_data = b.get("data", {}) if isinstance(b.get("data"), dict) else {}
            eff_name = eff_data.get("name", "无")
            eff_desc = eff_data.get("desc", "无")
            eff_tags = " ".join(eff_data.get("list", []))

            eff_display = f"{eff_name} ({eff_tags})" if eff_tags else eff_name
            if eff_name == "无":
                eff_display = "-"
                eff_desc = "-"

            item_idx = NumericTableWidgetItem(f"{idx}层", idx)
            item_name = QTableWidgetItem(name)
            item_skill = QTableWidgetItem(skills)
            item_eff = QTableWidgetItem(eff_display)
            item_desc = QTableWidgetItem(eff_desc)

            # Highlight special effect rows
            if eff_name != "无" and eff_name != "-":
                gold_brush = QBrush(QColor("#ffd54f"))
                item_eff.setForeground(gold_brush)
                item_eff.setFont(QFont("Microsoft YaHei UI", 9, QFont.Weight.Bold))

            item_idx.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_name.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            self.table_roster.setItem(i, 0, item_idx)
            self.table_roster.setItem(i, 1, item_name)
            self.table_roster.setItem(i, 2, item_skill)
            self.table_roster.setItem(i, 3, item_eff)
            self.table_roster.setItem(i, 4, item_desc)

        self.table_roster.setSortingEnabled(True)

    def export_json(self):
        try:
            p = self.mgr.export_json()
            bench_mgr = getattr(self.mgr, "bench_mgr", None)
            bench_cnt = bench_mgr.count() if bench_mgr else 0
            tip_extra = f"\n（已排除待选区 {bench_cnt} 个角色）" if bench_cnt > 0 else ""
            QMessageBox.information(self, "导出成功", f"JSON 文件已导出至:\n{p}{tip_extra}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出 JSON 失败: {e}")

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出 CSV 文件", "export.csv", "CSV Files (*.csv)")
        if not path:
            return
        try:
            active_chars = self.mgr.get_active_characters().values()
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(["角色", "区服", "门派", "等级", "装备分", "资历", "金币", "休闲点", "侠义", "百战精耐", "本地数据时间"])
                for c in active_chars:
                    bz_api = c.get("baizhan_api", {}) or {}
                    stamina = bz_api.get("skillStamina")
                    energy = bz_api.get("skillEnergy")
                    bz_jn_str = f"{stamina}/{energy}" if stamina is not None and energy is not None else "-"
                    w.writerow([
                        c.get("name", ""),
                        f"{c.get('region', '')}/{c.get('server', '')}",
                        c.get("force_name", ""),
                        c.get("level", 0),
                        c.get("equip_score", 0),
                        c.get("achievement_score", 0),
                        c.get("gold", 0),
                        c.get("contribution", 0),
                        c.get("justice", 0),
                        bz_jn_str,
                        c.get("last_update", "未知")
                    ])
            bench_mgr = getattr(self.mgr, "bench_mgr", None)
            bench_cnt = bench_mgr.count() if bench_mgr else 0
            tip_extra = f"\n（已排除待选区 {bench_cnt} 个角色）" if bench_cnt > 0 else ""
            QMessageBox.information(self, "导出成功", f"CSV 文件已导出至:\n{path}{tip_extra}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出 CSV 失败: {e}")

    def open_api_config_dialog(self):
        config = get_cached_config()
        dlg = ApiConfigDialog(config, self)
        if dlg.exec() == QDialog.DialogCode.Accepted and getattr(dlg, "token_modified", False):
            self.statusBar().showMessage("✓ API Token 已更新", 4000)


def main():
    app = QApplication(sys.argv)
    icon = get_app_icon()
    if not icon.isNull():
        app.setWindowIcon(icon)
    
    # Configuration Check & Auto Detection / Self-Healing
    config = get_cached_config()
    current_gp = config.get("game_path", "")
    auto_detect_msg = None

    # 1. 若路径未配置、不存在、或存在但缺少 my#data（用户填浅/游戏搬家），尝试自动探测与自愈
    if not current_gp or not os.path.exists(current_gp) or not is_valid_game_path(current_gp):
        detected_path, source = detect_game_path(current_gp if current_gp else None)
        if detected_path:
            config["game_path"] = detected_path
            save_config(config)
            auto_detect_msg = f"已自动检测并配置游戏路径: {detected_path} (来源: {source})"
            logger.info(auto_detect_msg)

    # 2. 校验配置；若仍不满足（如 api_key 缺失或未检测到有效路径）则弹出配置引导弹窗
    if validate_config(config) or not is_valid_game_path(config.get("game_path", "")):
        dlg = ConfigDialog(config)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            sys.exit(0)
        config = get_cached_config()
            
    mgr = JX3Manager()
    win = MainWindow(mgr)
    if auto_detect_msg:
        win.statusBar().showMessage(auto_detect_msg, 6000)
    win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
